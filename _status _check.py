import openpyxl
import os
import re

def resource_path(relative):
    base = os.path.dirname(__file__)
    return os.path.join(base, relative)

# Path to the output file
output_file = resource_path("Excel Files/Output.xlsx")
# Path to the input file
input_file = resource_path("Excel Files/Input.xlsx")

# Load the output workbook and select the Provider sheet
wb = openpyxl.load_workbook(output_file)
ws = wb['Provider']

# Load the input workbook and select the active sheet
wb_in = openpyxl.load_workbook(input_file)
ws_in = wb_in.active

# Get all rows as a list
rows = list(ws.iter_rows(values_only=True))

if not rows:
    print("No data found in the output file.")
    exit()

header = rows[0]
col_groups = {}  # base name -> list of (col_name, col_idx)

# Group columns by base name (e.g., 'Specialty 1', 'Specialty 2' -> 'Specialty')
for col_idx, col_name in enumerate(header):
    col_name_str = str(col_name) if col_name is not None else ""
    base = re.sub(r'\s*\d+$', '', col_name_str)
    if base not in col_groups:
        col_groups[base] = []
    col_groups[base].append((col_name_str, col_idx))

# List of columns to exclude if they are empty but have dropdowns (data validation)
dropdown_columns = {'Opt Out of Ratings', 'Enterprise Scheduling Flag'}

# Track columns whose header ends with 'ID' for a separate table
id_columns = []

# For each group, check if all columns in the group are completely empty
empty_groups = {}
for base, cols in col_groups.items():
    # If any column in the group has at least one value, skip this group
    group_has_value = False
    for col_name, col_idx in cols:
        if any((row[col_idx] is not None and row[col_idx] != "") for row in rows[1:]):
            group_has_value = True
            break
    # If the column header ends with 'ID', add to id_columns and skip main table
    if base.strip().endswith('ID'):
        id_columns.append((base, cols[0][1]))
        continue
    # Exclude columns with dropdowns
    if not group_has_value and base not in dropdown_columns:
        # Store the first column index for table display
        empty_groups[base] = cols[0][1]

# Function to convert column index (0-based) to Excel column letters
def col_idx_to_excel_letters(idx):
    letters = ''
    while idx >= 0:
        letters = chr(idx % 26 + ord('A')) + letters
        idx = idx // 26 - 1
    return letters

# Calculate statistics
if ws_in is not None:
    input_header_row = next(ws_in.iter_rows(min_row=1, max_row=1), [])
else:
    input_header_row = []
num_input_columns = len([cell for cell in input_header_row if cell.value is not None])
num_output_columns = len(header)
num_columns_left = len(empty_groups)
num_columns_done = num_output_columns - num_columns_left

print(f"Number of columns in Raw table (Input table): {num_input_columns}")
print(f"Number of columns left to work on: {num_columns_left}")
print(f"Total number of columns in the output file: {num_output_columns}")
print(f"Number of columns done: {num_columns_done}")

if empty_groups:
    # Prepare table data
    table_data = [(base, col_idx_to_excel_letters(idx)) for base, idx in sorted(empty_groups.items(), key=lambda x: x[1])]
    # Find max widths for formatting
    col1_width = max(len("Column Header"), max(len(row[0]) for row in table_data))
    col2_width = max(len("Column Number"), max(len(row[1]) for row in table_data))
    # Print number of items
    #print(f"Number of items: {len(table_data)}")
    # Print table header
    print(f"{'Column Header'.ljust(col1_width)} | {'Column Number'.ljust(col2_width)}")
    print(f"{'-'*col1_width}-+-{'-'*col2_width}")
    # Print table rows
    for base, col_letter in table_data:
        print(f"{base.ljust(col1_width)} | {col_letter.ljust(col2_width)}")
else:
    print("All columns have at least one value.")

# Print ID Columns table if any
if id_columns:
    id_table_data = [(base, col_idx_to_excel_letters(idx)) for base, idx in sorted(id_columns, key=lambda x: x[1])]
    id_col1_width = max(len("Column Header"), max(len(row[0]) for row in id_table_data))
    id_col2_width = max(len("Column Number"), max(len(row[1]) for row in id_table_data))
    print("\nID Columns.")
    print(f"{'Column Header'.ljust(id_col1_width)} | {'Column Number'.ljust(id_col2_width)}")
    print(f"{'-'*id_col1_width}-+-{'-'*id_col2_width}")
    for base, col_letter in id_table_data:
        print(f"{base.ljust(id_col1_width)} | {col_letter.ljust(id_col2_width)}")
