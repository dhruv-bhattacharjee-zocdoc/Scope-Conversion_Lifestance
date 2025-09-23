import openpyxl
from openpyxl.styles import PatternFill
import os
import re

def normalize_suffix(s):
    return re.sub(r'[^a-zA-Z0-9]', '', s or '').lower()

def get_dropdown_suffixes(template_file):
    wb = openpyxl.load_workbook(template_file, data_only=True)
    ws = wb['ValidationAndReference']
    suffixes = [ws.cell(row=i, column=7).value for i in range(2, 512)]
    suffixes = [s for s in suffixes if s and str(s).strip()]
    return suffixes

def highlight_invalid_suffixes(merged_file, template_file):
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    dropdown_suffixes = get_dropdown_suffixes(template_file)
    norm_dropdown = {normalize_suffix(s) for s in dropdown_suffixes}

    wb = openpyxl.load_workbook(merged_file)
    ws = wb['Provider']
    header_row = [cell.value for cell in ws[1]]
    max_row = ws.max_row

    for col_name in [f'Professional Suffix {i}' for i in range(1, 4)]:
        try:
            col_idx = header_row.index(col_name) + 1
        except ValueError:
            continue
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value
            norm_val = normalize_suffix(val) if val else ''
            if val and norm_val not in norm_dropdown:
                cell.fill = yellow_fill
            else:
                cell.fill = PatternFill(fill_type=None)
    wb.save(merged_file)

if __name__ == "__main__":
    merged_file = os.path.join("Excel Files", "Mergedoutput.xlsx")
    template_file = os.path.join("Excel Files", "New Business Scope Sheet - Practice Locations and Providers.xlsx")
    highlight_invalid_suffixes(merged_file, template_file)
