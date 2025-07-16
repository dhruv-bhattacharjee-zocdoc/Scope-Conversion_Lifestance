import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation


def set_opt_out_of_ratings_dropdown(output_file):
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Provider']
    # Find the column index for 'Opt Out of Ratings'
    header_row = [cell.value for cell in ws[1]]
    try:
        col_idx = header_row.index('Opt Out of Ratings') + 1  # 1-based index for openpyxl
    except ValueError:
        raise ValueError("'Opt Out of Ratings' column not found in output file.")
    # Create the data validation dropdown
    dv = DataValidation(type="list", formula1='"Yes"', allow_blank=True)
    # The range should cover all rows in the column except the header
    from openpyxl.utils import get_column_letter
    max_row = ws.max_row
    col_letter = get_column_letter(col_idx)
    dv_range = f"{col_letter}2:{col_letter}{max_row}"
    dv.add(dv_range)
    ws.add_data_validation(dv)
    wb.save(output_file)


