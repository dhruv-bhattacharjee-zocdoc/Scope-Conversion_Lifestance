import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict

# === Step 1: Generate practicecheck.xlsx ===
practice_location_path = 'Excel Files/Practice-Location.xlsx'
practicecheck_path = 'Excel Files/practicecheck.xlsx'

try:
    full_df = pd.read_excel(practice_location_path)
    practice_check_df = full_df[['Practice ID', 'Practice Cloud ID']].drop_duplicates(subset=['Practice Cloud ID'])
    practice_check_df.to_excel(practicecheck_path, index=False, sheet_name='practicecheck')
    print(f"Successfully wrote {practicecheck_path}")
except Exception as e:
    print(f"Error creating {practicecheck_path}: {e}")
    exit(1)

# === Step 2: Compare and highlight mismatches in Mergedoutput.xlsx ===
mergedoutput_path = 'Excel Files/Mergedoutput.xlsx'
input_path = 'Excel Files/Input.xlsx'

# Load mappings
practice_cloudid_to_id = dict(zip(practice_check_df['Practice Cloud ID'], practice_check_df['Practice ID']))
input_df = pd.read_excel(input_path)
npi_to_practiceid = dict(zip(input_df['NPI'], input_df['Practice ID']))

wbook = load_workbook(mergedoutput_path)
ws = wbook['Provider']
headers = [cell.value for cell in ws[1]]
id_idx = headers.index('Practice Cloud ID') + 1
npi_idx = headers.index('NPI Number') + 1

red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    practice_cloud_id = row[id_idx-1].value
    npi_number = row[npi_idx-1].value
    first_number = practice_cloudid_to_id.get(practice_cloud_id, None)
    second_number = npi_to_practiceid.get(npi_number, None)
    if first_number is not None and second_number is not None and str(first_number) != str(second_number):
        row[id_idx-1].fill = red_fill

wbook.save(mergedoutput_path)
print("Processed and highlighted mismatches in Mergedoutput.xlsx.")

# =============== Step 3: Location Type Consistency Check and Highlighting ===============

# Load Practice-Location.xlsx
practice_location_df = pd.read_excel(practice_location_path)

# Build a mapping from 'location_id' to 'Location Type' (from Practice-Location.xlsx)
pl_id_to_type = dict(zip(practice_location_df['location_id'], practice_location_df['Location Type']))

# Load Location tab from mergedoutput
location_ws = wbook['Location']
loc_headers = [cell.value for cell in location_ws[1]]
loc_id_col = loc_headers.index('Location Cloud ID') + 1
loc_type_col = loc_headers.index('Location Type') + 1

# Map 'Location Cloud ID' to row index in the Location tab (for highlighting)
locid_to_row = {}
locid_to_type = {}
for i, row in enumerate(location_ws.iter_rows(min_row=2, max_row=location_ws.max_row), start=2):
    loc_id = row[loc_id_col-1].value
    loc_type = row[loc_type_col-1].value
    if loc_id is not None:
        locid_to_row[str(loc_id).strip()] = i
        locid_to_type[str(loc_id).strip()] = loc_type

# Get column indices for Provider tab
provider_locid1_idx = headers.index('Location ID 1') + 1
provider_locid2_idx = headers.index('Location ID 2') + 1

# For each 'Location ID 1' and 'Location ID 2', perform lookup and compare
for ridx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    for idx in [provider_locid1_idx, provider_locid2_idx]:
        location_id = row[idx-1].value
        if not location_id:
            continue
        location_id_str = str(location_id).strip()
        # --- type_1: from Location tab ---
        type_1 = locid_to_type.get(location_id_str)
        # --- type_2: from Practice-Location.xlsx ---
        type_2 = pl_id_to_type.get(location_id_str)
        if type_1 is not None and type_2 is not None and str(type_1).strip().lower() != str(type_2).strip().lower():
            # Highlight Provider tab cell
            ws.cell(row=ridx, column=idx).fill = red_fill
            # Highlight Location tab cell
            if location_id_str in locid_to_row:
                location_row_idx = locid_to_row[location_id_str]
                location_ws.cell(row=location_row_idx, column=loc_id_col).fill = red_fill

wbook.save(mergedoutput_path)
print("Location type consistency check done and mismatches highlighted.")
