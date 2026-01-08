import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Input/output filenames
INPUT_FILE = 'Excel Files/Mergedoutput.xlsx'
OUTPUT_FILE = 'Excel Files/Report.xlsx'   # changed path here
SHEET_NAME = 'Provider'

# Color mappings (normalized RGB hex without #, as openpyxl stores "RRGGBB")
YELLOW_FILL = 'FFFF99'  # Professional Suffix 1-3 highlight
BLUE_FILL = '9BD7FF'    # NPI highlight
BRIGHT_YELLOW = 'FFFF00' # Professional Statement (>2000 chars)
RED = 'FF0000'           # Practice Cloud ID/Location ID anomaly

HEADER_FILL = PatternFill(start_color='4A206A', end_color='4A206A', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_ALIGNMENT = Alignment(horizontal='center')

SECONDROW_FONT = Font(italic=True, color='000000')
SECONDROW_ALIGNMENT = Alignment(wrap_text=True, horizontal='center', vertical='center')

THIN_BORDER = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

# Columns of interest
PS_SUFFIX_COLS = ['Professional Suffix 1','Professional Suffix 2','Professional Suffix 3']
SPECIALTY1_COL = 'Specialty 1'
NPI_COL = 'NPI Number'
PROF_STATEMENT_COL = 'Professional Statement'
PRACTICE_ID_COL = 'Practice Cloud ID'
LOCATION_ID_COLS = ['Location ID 1', 'Location ID 2']

# Category headers (user-edited)
CATEGORY_HEADERS = [
    'Professional Suffix Review',
    'Duplicate NPIs',
    'Professional Statement >2000)',
    'Practice Cloud ID Review',
    'Location Cloud ID Review'
]

# Second row descriptions (one above each column header)
CATEGORY_DESCRIPTIONS = [
    "Columns require manual editing of the suffix, as the scope did not clearly specify the suffix.",
    "Duplicate NPIs are highlighted in light blue. Manual merging is required.",
    "PFS contains more than 2,000 characters, which violates PFS regulations. Manual review is required.",
    "The Practice Cloud ID extracted and mapped does not match the requested one. Manual extraction is suggested.",
    "The Location ID is mapped incorrectly. Check the location type or whether locations are built under a different practice."
]

def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb[SHEET_NAME]
    header = [cell.value for cell in ws[1]]
    col_idx = {col: header.index(col)+1 for col in header if col is not None}
    
    yellow_rows = set()
    for col_name in PS_SUFFIX_COLS:
        if col_name in col_idx:
            col = col_idx[col_name]
            for row in range(2, ws.max_row+1):
                cell = ws.cell(row=row, column=col)
                fill = cell.fill
                if fill and fill.patternType and (get_hex(fill.fgColor.rgb) == YELLOW_FILL or get_hex(fill.start_color.rgb) == YELLOW_FILL):
                    yellow_rows.add(row)

    npi_blue_rows = []
    if NPI_COL in col_idx:
        col = col_idx[NPI_COL]
        for row in range(2, ws.max_row+1):
            cell = ws.cell(row=row, column=col)
            fill = cell.fill
            if fill and fill.patternType and (get_hex(fill.fgColor.rgb) == BLUE_FILL or get_hex(fill.start_color.rgb) == BLUE_FILL):
                npi_blue_rows.append(row)

    pfs_long_or_yellow = []
    if PROF_STATEMENT_COL in col_idx:
        col = col_idx[PROF_STATEMENT_COL]
        for row in range(2, ws.max_row+1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val and isinstance(val, str) and len(val) > 2000:
                pfs_long_or_yellow.append(row)
            fill = cell.fill
            if fill and fill.patternType and (get_hex(fill.fgColor.rgb) == BRIGHT_YELLOW or get_hex(fill.start_color.rgb) == BRIGHT_YELLOW):
                if row not in pfs_long_or_yellow:
                    pfs_long_or_yellow.append(row)
    
    practice_red_rows = []
    if PRACTICE_ID_COL in col_idx:
        col = col_idx[PRACTICE_ID_COL]
        for row in range(2, ws.max_row+1):
            cell = ws.cell(row=row, column=col)
            fill = cell.fill
            if fill and fill.patternType and (get_hex(fill.fgColor.rgb) == RED or get_hex(fill.start_color.rgb) == RED):
                practice_red_rows.append(row)
    
    location_red_rows = set()
    for col_name in LOCATION_ID_COLS:
        if col_name in col_idx:
            col = col_idx[col_name]
            for row in range(2, ws.max_row+1):
                cell = ws.cell(row=row, column=col)
                fill = cell.fill
                if fill and fill.patternType and (get_hex(fill.fgColor.rgb) == RED or get_hex(fill.start_color.rgb) == RED):
                    location_red_rows.add(row)
    
    # Existing finding columns
    write_data = [
        sorted(yellow_rows), 
        npi_blue_rows, 
        pfs_long_or_yellow, 
        practice_red_rows, 
        sorted(location_red_rows)
    ]

    # Add green-highlighted rows for 'Telehealth' Location ID 2s
    TELEHEALTH_GREEN = '00FF00'
    telehealth_green_rows = []
    if 'Location ID 2' in col_idx:
        col = col_idx['Location ID 2']
        for row in range(2, ws.max_row+1):
            cell = ws.cell(row=row, column=col)
            fill = cell.fill
            if fill and fill.patternType and (get_hex(fill.fgColor.rgb) == TELEHEALTH_GREEN or get_hex(fill.start_color.rgb) == TELEHEALTH_GREEN):
                telehealth_green_rows.append(row)
    write_data.append(telehealth_green_rows)

    write_report(write_data)
    # --- Call sheetmerge.py as the last step ---
    import subprocess
    subprocess.run(['python', 'sheetmerge.py'], check=True)

# Add a header and description for the new column
CATEGORY_HEADERS.append('Telehealth Green Highlighted Rows')
CATEGORY_DESCRIPTIONS.append('Provider Requested for Telehealth. Review and Correct Mapping.')

def get_hex(rgb):
    """Normalize openpyxl ARGB or RGB formats to 'RRGGBB'."""
    if rgb is None:
        return None
    if len(rgb) == 8:
        return rgb[2:]
    if len(rgb) == 6:
        return rgb.upper()
    return rgb

def write_report(data_columns):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Provider'
    num_cats = len(CATEGORY_HEADERS)
    max_height = max(len(data) for data in data_columns)
    # Insert two rows at the top
    ws.insert_rows(1, 2)
    # Write the second row - descriptions (row 2)
    for idx, desc in enumerate(CATEGORY_DESCRIPTIONS):
        col = 1 + idx * 2
        cell = ws.cell(row=2, column=col)
        cell.value = desc
        cell.font = SECONDROW_FONT
        cell.alignment = SECONDROW_ALIGNMENT
        cell.border = THIN_BORDER
    # Write the third row - headers (row 3)
    for idx, header in enumerate(CATEGORY_HEADERS):
        col = 1 + idx * 2
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    # Fill in data starting row 4
    for idx, datalist in enumerate(data_columns):
        col = 1 + idx * 2
        for i, val in enumerate(datalist):
            c = ws.cell(row=i+4, column=col)
            c.value = val
            c.border = THIN_BORDER
    # Set column width to 200px (~26.6 points/characters for Excel columns)
    # 1 character is ~7.5px, so 200px/7.5=26.6 Excel width
    px_width = 200
    excel_width = 26.6
    for idx in range(num_cats):
        col = 1 + idx * 2
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = excel_width
    # All used cells get a border (except blank columns)
    for idx in range(num_cats):
        col = 1 + idx * 2
        for r in range(2, max_height+5):
            ws.cell(row=r, column=col).border = THIN_BORDER
    wb.save(OUTPUT_FILE)

if __name__ == '__main__':
    main()
