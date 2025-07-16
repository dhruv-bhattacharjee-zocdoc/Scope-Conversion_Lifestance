import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def extract_name_gender(input_file):
    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in.active
    if ws_in is None:
        raise ValueError("Input worksheet could not be loaded.")
    input_header_row = [cell.value for cell in ws_in[1]]
    input_indices = {header: idx for idx, header in enumerate(input_header_row) if header in ['First Name', 'Last Name', 'Gender']}
    extracted_rows = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        extracted = {col: row[input_indices[col]] if col in input_indices else None for col in ['First Name', 'Last Name', 'Gender']}
        extracted_rows.append(extracted)
    return extracted_rows


def add_gender_dropdown(ws, header_row=1):
    """
    Adds a gender dropdown (Male, Female, NonBinary, Not Applicable) to the 'Gender' column of the given worksheet.
    Does not save the workbook. Call this after writing data to the output sheet.
    """
    header = [cell.value for cell in ws[header_row]]
    try:
        gender_idx = header.index('Gender')
    except ValueError:
        return  # Gender column not found
    gender_col_letter = get_column_letter(gender_idx+1)
    dv_gender = DataValidation(type="list", formula1='"Male,Female,NonBinary,Not Applicable"', allow_blank=True)
    dv_gender_range = f"{gender_col_letter}{header_row+1}:{gender_col_letter}{ws.max_row}"
    dv_gender.add(dv_gender_range)
    ws.add_data_validation(dv_gender)
