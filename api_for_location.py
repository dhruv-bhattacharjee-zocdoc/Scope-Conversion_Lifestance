from PracticeIDlist import extract_unique_practice_ids
import os
import openpyxl
import requests
import sys
import time
from openpyxl.styles import PatternFill, Font

# Step 1: Ensure Practice-Location.xlsx is created/updated
input_path = os.path.join("Excel Files", "Input.xlsx")
output_path = os.path.join("Excel Files", "Practice-Location.xlsx")
extract_unique_practice_ids(input_path, output_path)

# Step 2: Read unique Practice IDs from Practice-Location.xlsx
wb = openpyxl.load_workbook(output_path, data_only=True)
sheet = wb.active
practice_ids = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    pid = row[0]
    if pid is not None:
        practice_ids.append(str(pid))
unique_practice_ids = list(set(practice_ids))
print(f"Number of Unique Practice IDs in the Batch: {len(unique_practice_ids)}")

# Step 3: Make POST request to get cloud IDs
url = 'https://provider-reference-v1.east.zocdoccloud.com/provider-reference/v1/practice/ids-by-monolith-ids~batchGet'
headers = {
    'accept': 'application/json',
    'Content-Type': 'application/json'
}
data = {
    "monolith_practice_ids": unique_practice_ids
}
print("Fetching Practice Cloud IDs...")
response = requests.post(url, headers=headers, json=data)

cloud_id_map = {}
if response.status_code == 200:
    result = response.json()
    for item in result.get('practice_ids', []):
        monolith_id = str(item.get('monolith_practice_id'))
        cloud_id = item.get('practice_id')
        cloud_id_map[monolith_id] = cloud_id
else:
    print(f"Failed to fetch cloud IDs. Status code: {response.status_code}")
    print(response.text)
    cloud_id_map = {}

# Step 4: Write 'Practice Cloud ID' to Practice-Location.xlsx
wb = openpyxl.load_workbook(output_path)
sheet = wb.active
header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
if 'Practice Cloud ID' in header:
    cloud_col_idx = header.index('Practice Cloud ID') + 1
else:
    cloud_col_idx = len(header) + 1
    sheet.cell(row=1, column=cloud_col_idx, value='Practice Cloud ID')
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
    pid = str(row[0].value)
    cloud_id = cloud_id_map.get(pid, None)
    sheet.cell(row=row_idx, column=cloud_col_idx, value=cloud_id)
wb.save(output_path)
print(f"Updated {output_path} with Practice Cloud IDs.")

# Step 5: For each Practice Cloud ID, get location details and append to Excel
location_url = 'https://provider-reference-v1.east.zocdoccloud.com/provider-reference/v1/practice/location~batchGet'
location_headers = {
    'accept': 'application/json',
    'Content-Type': 'application/json'
}
# Define the fields to extract, with 'is_virtual' before 'address_1' and 'Location Type' after 'is_virtual'
location_fields = [
    'is_virtual', 'Location Type', 'address_1', 'address_2', 'city', 'state', 'zip',
    'monolith_location_id', 'location_id', 'virtual_visit_type',
    'software', 'software_id', 'hide_on_profile', 'phone', 'email_addresses'
]
wb = openpyxl.load_workbook(output_path)
sheet = wb.active
header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
for field in location_fields:
    if field not in header:
        header.append(field)
        sheet.cell(row=1, column=len(header), value=field)

# Apply header styling: fill color '000AD6', font color 'FFFFFF'
header_fill = PatternFill(start_color='000AD6', end_color='000AD6', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
for col in range(1, len(header) + 1):
    cell = sheet.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
# Freeze the header row
sheet.freeze_panes = 'A2'
# Add autofilter to all columns
sheet.auto_filter.ref = sheet.dimensions

practice_id_to_cloud_id = {}
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    pid = row[0].value
    cloud_id = row[cloud_col_idx-1].value
    if pid and cloud_id:
        practice_id_to_cloud_id[str(cloud_id)] = str(pid)
cloud_ids = list(practice_id_to_cloud_id.keys())
print(f"Number of Practice IDs to process for location details: {len(cloud_ids)}")
processed_count = 0
for cloud_id in cloud_ids:
    data = {"practice_ids": [cloud_id]}
    response = requests.post(location_url, headers=location_headers, json=data)
    if response.status_code == 200:
        result = response.json()
        locations = result.get('practice_locations', [])
        for loc in locations:
            row_data = [practice_id_to_cloud_id[cloud_id], cloud_id]
            # is_virtual
            is_virtual_value = loc.get('is_virtual', None)
            row_data.append(is_virtual_value)
            # Location Type
            if is_virtual_value is True or (isinstance(is_virtual_value, str) and is_virtual_value.upper() == 'TRUE'):
                row_data.append('Virtual')
            elif is_virtual_value is False or (isinstance(is_virtual_value, str) and is_virtual_value.upper() == 'FALSE'):
                row_data.append('In Person')
            else:
                row_data.append(None)
            # The rest of the fields
            for field in location_fields[2:]:
                value = loc.get(field, None)
                if isinstance(value, list):
                    value = ', '.join(map(str, value))
                row_data.append(value)
            sheet.append(row_data)
    else:
        print(f"Failed to fetch locations for Cloud ID {cloud_id}. Status code: {response.status_code}")
        print(response.text)
    processed_count += 1
    sys.stdout.write(f"\rProcessed Practice IDs: {processed_count}/{len(cloud_ids)} ")
    sys.stdout.flush()
print()  # Move to next line after progress

# Step 6: Remove rows with only Practice ID and Practice Cloud ID before saving
rows_to_delete = []
header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
cloud_col_idx = header.index('Practice Cloud ID') + 1
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
    values = [cell.value for cell in row]
    # Only Practice ID and Practice Cloud ID have values, rest are None or empty
    if all((v is None or v == '') for v in values[cloud_col_idx:]):
        rows_to_delete.append(row_idx)
for idx in reversed(rows_to_delete):
    sheet.delete_rows(idx)

wb.save(output_path)
print(f"Appended location details to {output_path} and cleaned up empty rows.")
