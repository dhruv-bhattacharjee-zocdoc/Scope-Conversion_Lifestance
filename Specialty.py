import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def extract_specialty(input_file):
    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in.active
    if ws_in is None:
        raise ValueError("Input worksheet could not be loaded.")
    input_header_row = [cell.value for cell in ws_in[1]]
    try:
        specialty_idx = input_header_row.index('Board Subspecialty')
    except ValueError:
        raise ValueError("'Board Subspecialty' column not found in input file.")
    specialty_list = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        specialty_list.append(row[specialty_idx])
    return specialty_list


def add_specialty_dropdowns(ws, ws_valref, header_row=1):
    """
    Adds dropdowns to 'Specialty 1', 'Specialty 2', 'Specialty 3', ... columns in ws,
    using unique, non-empty values from the 'Specialty Name' column in ws_valref (ValidationAndReference sheet).
    Does not save the workbook.
    """
    valref_header = [cell.value for cell in ws_valref[header_row]]
    try:
        specialty_name_idx = valref_header.index('Specialty Name')
    except ValueError:
        return  # Specialty Name column not found
    # Collect unique, non-empty specialty values
    specialty_values = set()
    for row in ws_valref.iter_rows(min_row=header_row+1, max_row=ws_valref.max_row, values_only=True):
        val = row[specialty_name_idx] if specialty_name_idx < len(row) else None
        if val and str(val).strip():
            specialty_values.add(str(val).strip())
    specialty_list = sorted(specialty_values)
    if not specialty_list:
        return
    # Find all Specialty columns
    ws_header = [cell.value for cell in ws[header_row]]
    specialty_cols = []
    for idx, col_name in enumerate(ws_header):
        if isinstance(col_name, str) and col_name.startswith('Specialty'):
            specialty_cols.append(idx)
    # Add dropdowns to each column
    for idx in specialty_cols:
        col_letter = get_column_letter(idx+1)
        dv = DataValidation(type="list", formula1='"' + ','.join(specialty_list) + '"', allow_blank=True)
        dv_range = f"{col_letter}{header_row+1}:{col_letter}{ws.max_row}"
        dv.add(dv_range)
        ws.add_data_validation(dv)
