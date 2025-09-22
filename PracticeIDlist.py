import openpyxl
from openpyxl import Workbook
import os

def extract_unique_practice_ids(input_excel_path, output_excel_path):
    # Load the workbook and select the first sheet
    wb = openpyxl.load_workbook(input_excel_path, data_only=True)
    sheet = wb.active

    # Find the column index for 'Practice ID'
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    try:
        practice_id_col_idx = header.index('Practice ID')
    except ValueError:
        raise Exception("'Practice ID' column not found in the Excel file.")

    # Extract all unique values from the 'Practice ID' column (skip header)
    practice_id_set = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        practice_id_value = row[practice_id_col_idx]
        if practice_id_value is not None:
            practice_id_set.add(practice_id_value)

    # Write the unique Practice IDs to a new Excel file
    wb_out = Workbook()
    sheet_out = wb_out.active
    sheet_out.title = 'Practice ID'
    # Write header
    sheet_out['A1'] = 'Practice ID'
    # Write Practice ID values (ignore text types, only keep int/float)
    numeric_ids = [pid for pid in practice_id_set if isinstance(pid, (int, float))]
    for idx, pid in enumerate(sorted(numeric_ids), start=2):
        sheet_out[f'A{idx}'] = pid
    try:
        wb_out.save(output_excel_path)
        print(f"Extracted {len(practice_id_set)} unique Practice ID values and saved to {output_excel_path}")
    except Exception as e:
        print(f"Failed to save file at {output_excel_path}. Error: {e}")
        print(f"Current working directory: {os.getcwd()}")

if __name__ == "__main__":
    input_path = os.path.join("Excel Files", "Input.xlsx")
    output_path = os.path.join("Excel Files", "Practice-Location.xlsx")
    extract_unique_practice_ids(input_path, output_path)