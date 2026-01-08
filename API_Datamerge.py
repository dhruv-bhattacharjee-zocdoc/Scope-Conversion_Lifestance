import shutil
import subprocess
import os
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from rapidfuzz import process, fuzz
import re
import sys
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

# Step 1: Copy Output.xlsx to Mergedoutput.xlsx
src = r"Excel Files/Output.xlsx"
dst = r"Excel Files/Mergedoutput.xlsx"
shutil.copyfile(src, dst)
print(f"Copied {src} to {dst}")

# Step 2: Run api_for_specialty.py
print("Running api_for_specialty.py...")
subprocess.run(["python", "api_for_specialty.py"], check=True)

# Step 3: Run api_for_location.py
print("Running api_for_location.py...")
subprocess.run(["python", "api_for_location.py"], check=True)

# Step 3.5: Update Location sheet in Mergedoutput.xlsx using Practice-Location.xlsx
practice_location_path = r"Excel Files/Practice-Location.xlsx"

# Read both sheets as DataFrames
merged_file_path = os.path.abspath(dst)
loc_df = pd.read_excel(merged_file_path, sheet_name='Location')
prac_df = pd.read_excel(practice_location_path)

# Ensure ZIP Code columns are always 5-character strings with leading zeros
if 'ZIP Code' in loc_df.columns:
    loc_df['ZIP Code'] = loc_df['ZIP Code'].astype(str).str.zfill(5)
if 'zip' in prac_df.columns:
    prac_df['zip'] = prac_df['zip'].astype(str).str.zfill(5)

# Define matching columns
loc_match_cols = ['Address line 1', 'Address line 2 (Office/Suite #)', 'Location Type', 'City', 'State', 'ZIP Code']
prac_match_cols = ['address_1', 'address_2', 'Location Type', 'city', 'state', 'zip']

# Fill NaN for address_2 if missing in either
loc_df['Address line 2 (Office/Suite #)'] = loc_df['Address line 2 (Office/Suite #)'].fillna("")
prac_df['address_2'] = prac_df['address_2'].fillna("")

# Ensure relevant columns are string type to avoid dtype warnings
for col in [
    'Practice Cloud ID', 'Location Cloud ID', 'Scheduling Software',
    'Scheduling Software ID', 'Phone', 'Virtual Visit Type',
    'Email for appointment notifications 1', 'Practice Name', 'Location Name'
]:
    if col in loc_df.columns:
        loc_df[col] = loc_df[col].astype(str)

def try_fuzzy_with_address2(sub_prac_df, loc_row, addr2_col, loc_addr2):
    choices = sub_prac_df['address_1'].tolist()
    best = process.extractOne(loc_row['Address line 1'], choices, scorer=fuzz.token_sort_ratio)
    if best and best[1] > 90:
        best_row = sub_prac_df[sub_prac_df['address_1'] == best[0]].iloc[0]
        return best_row
    partial_candidates = process.extract(loc_row['Address line 1'], choices, scorer=fuzz.token_sort_ratio, score_cutoff=70)
    for cand_addr1, score, idx_in_sub in partial_candidates:
        candidate = sub_prac_df.iloc[idx_in_sub]
        candidate_addr2 = candidate[addr2_col]
        if pd.isna(loc_addr2) and pd.isna(candidate_addr2):
            return candidate
        candidate_addr2 = str(candidate_addr2) if not pd.isna(candidate_addr2) else ""
        loc_addr2_str = str(loc_addr2) if not pd.isna(loc_addr2) else ""
        if candidate_addr2.strip() == loc_addr2_str.strip():
            return candidate
        if (candidate_addr2.strip() != "" and loc_addr2_str.strip() != ""):
            fuzz_score = fuzz.token_sort_ratio(candidate_addr2, loc_addr2_str)
            if fuzz_score > 90:
                return candidate
    return None

for idx, loc_row in loc_df.iterrows():
    addr2 = loc_row['Address line 2 (Office/Suite #)']
    best_row = None
    if not addr2 or str(addr2).strip() == '':
        # Only allow matches in prac_df where address_2 is also blank
        sub_prac_df = prac_df[
            (prac_df['Location Type'] == loc_row['Location Type']) &
            (prac_df['city'] == loc_row['City']) &
            (prac_df['state'] == loc_row['State']) &
            (prac_df['zip'] == loc_row['ZIP Code']) &
            ((prac_df['address_2'].isnull()) | (prac_df['address_2'].astype(str).str.strip() == ''))
        ]
        match = sub_prac_df[
            (sub_prac_df['address_1'] == loc_row['Address line 1'])
        ]
        if not match.empty:
            best_row = match.iloc[0]
        else:
            best_row = try_fuzzy_with_address2(sub_prac_df, loc_row, 'address_2', loc_row['Address line 2 (Office/Suite #)'])
    else:
        sub_prac_df = prac_df[
            (prac_df['address_2'] == addr2) &
            (prac_df['Location Type'] == loc_row['Location Type']) &
            (prac_df['city'] == loc_row['City']) &
            (prac_df['state'] == loc_row['State']) &
            (prac_df['zip'] == loc_row['ZIP Code'])
        ]
        match = sub_prac_df[
            (sub_prac_df['address_1'] == loc_row['Address line 1'])
        ]
        if not match.empty:
            best_row = match.iloc[0]
        else:
            best_row = try_fuzzy_with_address2(sub_prac_df, loc_row, 'address_2', addr2)
    if best_row is None:
        # FINAL fallback: fuzzy match on address_1 only across all practice locations
        all_choices = prac_df['address_1'].tolist()
        best_address = process.extractOne(loc_row['Address line 1'], all_choices, scorer=fuzz.token_sort_ratio)
        if best_address:
            best_row = prac_df[prac_df['address_1'] == best_address[0]].iloc[0]
    if best_row is not None:
        loc_df.at[idx, 'Practice Cloud ID'] = str(best_row.get('Practice Cloud ID', ''))
        loc_df.at[idx, 'Location Cloud ID'] = str(best_row.get('location_id', ''))
        loc_df.at[idx, 'Scheduling Software'] = str(best_row.get('software', ''))
        loc_df.at[idx, 'Scheduling Software ID'] = str(best_row.get('software_id', ''))
        loc_df.at[idx, 'Phone'] = str(best_row.get('phone', ''))
        loc_df.at[idx, 'Virtual Visit Type'] = str(best_row.get('virtual_visit_type', ''))
        loc_df.at[idx, 'Email for appointment notifications 1'] = 'practice.manager@zocdocusername.com'
        loc_df.at[idx, 'Practice Name'] = 'LifeStance Health'
        loc_df.at[idx, 'Location Name'] = 'LifeStance Health'

# Write the updated Location sheet back to the workbook
with pd.ExcelWriter(merged_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    loc_df.to_excel(writer, sheet_name='Location', index=False)

print("Updated Location sheet in Mergedoutput.xlsx using Practice-Location.xlsx.")

# --- NEW: Re-map columns 'Location ID 1' through 'Location ID 5' in the Provider sheet from updated Location sheet ---
wb = openpyxl.load_workbook(merged_file_path)
ws_provider = wb['Provider']
ws_location = wb['Location']

prov_header = [cell.value for cell in ws_provider[1]]
loc_header = [cell.value for cell in ws_location[1]]

location_cloud_id_idx = loc_header.index("Location Cloud ID") + 1
complete_location_idx = loc_header.index("Complete Location") + 1 if "Complete Location" in loc_header else None

for n in range(1, 6):
    try:
        col_idx = prov_header.index(f'Location ID {n}') + 1
    except ValueError:
        continue
    for row in range(2, ws_provider.max_row + 1):
        loc_id_val = ws_provider.cell(row=row, column=col_idx).value
        if loc_id_val:
            # Look up in Location sheet
            found = False
            for lrow in range(2, ws_location.max_row + 1):
                if ws_location.cell(row=lrow, column=location_cloud_id_idx).value == loc_id_val:
                    if complete_location_idx:
                        provider_loc_col_name = f'Location {n}'
                        try:
                            provider_loc_col = prov_header.index(provider_loc_col_name) + 1
                            ws_provider.cell(row=row, column=provider_loc_col, value=ws_location.cell(row=lrow, column=complete_location_idx).value)
                        except ValueError:
                            pass  # If the column doesn't exist, skip
                    found = True
                    break
            if not found:
                # Write blank if no match found
                provider_loc_col_name = f'Location {n}'
                try:
                    provider_loc_col = prov_header.index(provider_loc_col_name) + 1
                    ws_provider.cell(row=row, column=provider_loc_col, value="")
                except ValueError:
                    pass
wb.save(merged_file_path)
print("Re-mapped Location ID 1-5 columns in Provider sheet from updated Location sheet.")

# Highlight unmatched rows in yellow (flexible matching for address_2)
wb_loc = openpyxl.load_workbook(merged_file_path)
ws_loc = wb_loc['Location']
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Re-read Practice-Location for matching
prac_df = pd.read_excel(practice_location_path)
prac_df['address_2'] = prac_df['address_2'].fillna("")

header_row = [cell.value for cell in ws_loc[1]]
for row in range(2, ws_loc.max_row + 1):
    addr2 = ws_loc.cell(row=row, column=header_row.index('Address line 2 (Office/Suite #)')+1).value
    if not addr2 or str(addr2).strip() == '':
        loc_vals = [
            ws_loc.cell(row=row, column=header_row.index('Address line 1')+1).value,
            ws_loc.cell(row=row, column=header_row.index('Location Type')+1).value,
            ws_loc.cell(row=row, column=header_row.index('City')+1).value,
            ws_loc.cell(row=row, column=header_row.index('State')+1).value,
            ws_loc.cell(row=row, column=header_row.index('ZIP Code')+1).value,
        ]
        match = prac_df[
            (prac_df['address_1'] == loc_vals[0]) &
            (prac_df['Location Type'] == loc_vals[1]) &
            (prac_df['city'] == loc_vals[2]) &
            (prac_df['state'] == loc_vals[3]) &
            (prac_df['zip'] == loc_vals[4])
        ]
    else:
        loc_vals = [
            ws_loc.cell(row=row, column=header_row.index('Address line 1')+1).value,
            addr2,
            ws_loc.cell(row=row, column=header_row.index('Location Type')+1).value,
            ws_loc.cell(row=row, column=header_row.index('City')+1).value,
            ws_loc.cell(row=row, column=header_row.index('State')+1).value,
            ws_loc.cell(row=row, column=header_row.index('ZIP Code')+1).value,
        ]
        match = prac_df[
            (prac_df['address_1'] == loc_vals[0]) &
            (prac_df['address_2'] == loc_vals[1]) &
            (prac_df['Location Type'] == loc_vals[2]) &
            (prac_df['city'] == loc_vals[3]) &
            (prac_df['state'] == loc_vals[4]) &
            (prac_df['zip'] == loc_vals[5])
        ]
    if match.empty:
        for col in range(1, ws_loc.max_column + 1):
            ws_loc.cell(row=row, column=col).fill = yellow_fill

wb_loc.save(merged_file_path)
print("Highlighted unmatched rows in Location sheet with yellow.")

print("Updated Location sheet in Mergedoutput.xlsx using Practice-Location.xlsx.")

# Insert formula in 'Complete Location' column before updating Location sheet
wb_loc_formula = openpyxl.load_workbook(merged_file_path)
ws_loc_formula = wb_loc_formula['Location']
header_row = [cell.value for cell in ws_loc_formula[1]]
try:
    complete_loc_col = header_row.index('Complete Location') + 1
except ValueError:
    raise Exception("'Complete Location' column not found in Location sheet.")
for row in range(2, ws_loc_formula.max_row + 1):
    practice_name_col = header_row.index('Practice Name') + 1
    practice_name_val = ws_loc_formula.cell(row=row, column=practice_name_col).value
    if str(practice_name_val).strip().lower() == 'nan':
        continue  # Skip this row
    formula = f'=IF(A{row}<>"",CONCATENATE(A{row}," ",B{row}," ",D{row}," ",E{row}," ",F{row}," ",G{row}," ","(",C{row},")"),"")'
    ws_loc_formula.cell(row=row, column=complete_loc_col, value=formula)
wb_loc_formula.save(merged_file_path)

# Step 4: Fill 'Specialty ID 1' in Provider sheet of Mergedoutput.xlsx
npi_specialty_path = r"Excel Files/Npi-specialty.xlsx"

# Load NPI-Specialty mapping (use correct column name 'SPECIALTIES')
npi_df = pd.read_excel(npi_specialty_path, dtype={"NPI": str})
npi_to_specialty = dict(zip(npi_df["NPI"].astype(str), npi_df["SPECIALTIES"]))

# Load Mergedoutput.xlsx and update Provider sheet
wb = openpyxl.load_workbook(merged_file_path)
ws = wb["Provider"]
header = [cell.value for cell in ws[1]]
try:
    npi_col = header.index("NPI Number") + 1
    specialty_col = header.index("Specialty ID 1") + 1
except ValueError as e:
    raise Exception(f"Required column not found: {e}")

# Fill 'Specialty ID 1'
for row in range(2, ws.max_row + 1):
    npi_value = ws.cell(row=row, column=npi_col).value
    if npi_value is not None:
        specialty = npi_to_specialty.get(str(npi_value))
        if specialty is not None:
            ws.cell(row=row, column=specialty_col, value=specialty)

# Set formula in 'Specialty 1' column
try:
    specialty1_col = header.index("Specialty 1") + 1
    for row in range(2, ws.max_row + 1):
        formula = f'=IFERROR(VLOOKUP(BM{row}, ValidationAndReference!$J:$K, 2, FALSE), "")'
        ws.cell(row=row, column=specialty1_col, value=formula)
except ValueError:
    print("'Specialty 1' column not found, skipping formula step.")

wb.save(merged_file_path)
print("Filled 'Specialty ID 1' and set formula in 'Specialty 1' in Provider sheet of Mergedoutput.xlsx.")

# Step 4: Run locationmapping.py as the last step
print("Running locationmapping.py as the last step...")
subprocess.run(["python", "locationmapping.py"], check=True)

# Insert formulas for 'Location 1' and 'Location 2' in Provider sheet
wb = openpyxl.load_workbook(merged_file_path)
ws = wb["Provider"]
header = [cell.value for cell in ws[1]]
try:
    location1_col = header.index("Location 1") + 1
    location2_col = header.index("Location 2") + 1
    for row in range(2, ws.max_row + 1):
        formula1 = f'=IFERROR(INDEX(Location!X:X, MATCH(BR{row}, Location!W:W, 0)), "")'
        formula2 = f'=IFERROR(INDEX(Location!X:X, MATCH(BS{row}, Location!W:W, 0)), "")'
        ws.cell(row=row, column=location1_col, value=formula1)
        ws.cell(row=row, column=location2_col, value=formula2)
    # Add dropdown validation for both columns
    from openpyxl.worksheet.datavalidation import DataValidation
    import openpyxl.utils
    dv = DataValidation(type="list", formula1="=Location!$X$2:$X$1000", allow_blank=True)
    ws.add_data_validation(dv)
    col1_range = f'{openpyxl.utils.get_column_letter(location1_col)}2:{openpyxl.utils.get_column_letter(location1_col)}{ws.max_row}'
    col2_range = f'{openpyxl.utils.get_column_letter(location2_col)}2:{openpyxl.utils.get_column_letter(location2_col)}{ws.max_row}'
    dv.add(col1_range)
    dv.add(col2_range)
    # Set formula for 'Specialty 1' column
    try:
        specialty1_col = header.index("Specialty 1") + 1
        for row in range(2, ws.max_row + 1):
            formula = f'=IFERROR(VLOOKUP(BM{row}, ValidationAndReference!J:K, 2, FALSE), "")'
            ws.cell(row=row, column=specialty1_col, value=formula)
        # Add dropdown validation for 'Specialty 1' through 'Specialty 5'
        dv_specialty = DataValidation(type="list", formula1="=ValidationAndReference!$K$2:$K$311", allow_blank=True)
        ws.add_data_validation(dv_specialty)
        for specialty_col_name in ["Specialty 1", "Specialty 2", "Specialty 3", "Specialty 4", "Specialty 5"]:
            try:
                col_idx = header.index(specialty_col_name) + 1
                specialty_range = f'{openpyxl.utils.get_column_letter(col_idx)}2:{openpyxl.utils.get_column_letter(col_idx)}{ws.max_row}'
                dv_specialty.add(specialty_range)
            except ValueError:
                print(f"'{specialty_col_name}' column not found, skipping validation for this column.")
    except ValueError:
        print("'Specialty 1' column not found, skipping formula step.")
    wb.save(merged_file_path)
    print("Inserted formulas and dropdown validation for 'Location 1', 'Location 2', and 'Specialty 1' in Provider sheet.")

    # Map 'Practice Cloud ID' and 'Practice Name' from Location sheet to Provider sheet using 'Location ID 1'
    try:
        location_id1_col = header.index("Location ID 1") + 1
        practice_cloud_id_col = header.index("Practice Cloud ID") + 1
        practice_name_col = header.index("Practice Name") + 1
        # Load Location sheet for lookup
        wb_loc = openpyxl.load_workbook(merged_file_path, data_only=True)
        ws_loc = wb_loc["Location"]
        loc_header = [cell.value for cell in ws_loc[1]]
        loc_cloud_id_idx = loc_header.index("Location Cloud ID") + 1
        practice_cloud_id_idx = loc_header.index("Practice Cloud ID") + 1
        practice_name_idx = loc_header.index("Practice Name") + 1
        for row in range(2, ws.max_row + 1):
            loc_id_1 = ws.cell(row=row, column=location_id1_col).value
            practice_cloud_id = ""
            practice_name = ""
            if loc_id_1:
                for loc_row in range(2, ws_loc.max_row + 1):
                    if ws_loc.cell(row=loc_row, column=loc_cloud_id_idx).value == loc_id_1:
                        practice_cloud_id = ws_loc.cell(row=loc_row, column=practice_cloud_id_idx).value
                        practice_name = ws_loc.cell(row=loc_row, column=practice_name_idx).value
                        break
            ws.cell(row=row, column=practice_cloud_id_col, value=practice_cloud_id)
            ws.cell(row=row, column=practice_name_col, value=practice_name)
        wb.save(merged_file_path)
        print("Mapped 'Practice Cloud ID' and 'Practice Name' from Location sheet to Provider sheet.")
    except ValueError as e:
        print(f"Required column not found for Practice Cloud ID or Practice Name mapping: {e}")
except ValueError as e:
    print(f"Required column not found for Location formulas: {e}")

# Add dropdown validation for 'Patients Accepted' column
try:
    patients_accepted_col = header.index('Patients Accepted') + 1
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    dv_patients = DataValidation(type="list", formula1='"Adult,Pediatric,Both"', allow_blank=True)
    col_letter = get_column_letter(patients_accepted_col)
    dv_range = f"{col_letter}2:{col_letter}{ws.max_row}"
    dv_patients.add(dv_range)
    ws.add_data_validation(dv_patients)
    wb.save(merged_file_path)
    print("Added dropdown validation for 'Patients Accepted' column in Provider sheet.")
except ValueError:
    print("'Patients Accepted' column not found, skipping dropdown validation.")

# Add dropdown validation for 'Gender' column in Provider sheet.
try:
    gender_col = header.index('Gender') + 1
    gender_col_letter = get_column_letter(gender_col)
    dv_gender = DataValidation(type="list", formula1='"Male,Female,NonBinary,Not Applicable"', allow_blank=True)
    dv_gender_range = f"{gender_col_letter}2:{gender_col_letter}{ws.max_row}"
    dv_gender.add(dv_gender_range)
    ws.add_data_validation(dv_gender)
    wb.save(merged_file_path)
    print("Added dropdown validation for 'Gender' column in Provider sheet.")
except ValueError:
    print("'Gender' column not found, skipping dropdown validation.")

# Add dropdown validation for 'Professional Suffix 1' through 'Professional Suffix 3' columns in Provider sheet.
for i in range(1, 4):
    col_name = f'Professional Suffix {i}'
    try:
        suffix_col = header.index(col_name) + 1
        suffix_col_letter = get_column_letter(suffix_col)
        dv_suffix = DataValidation(type="list", formula1='=ValidationAndReference!$G$2:$G$511', allow_blank=True)
        dv_suffix_range = f"{suffix_col_letter}2:{suffix_col_letter}{ws.max_row}"
        dv_suffix.add(dv_suffix_range)
        ws.add_data_validation(dv_suffix)
        wb.save(merged_file_path)
        print(f"Added dropdown validation for '{col_name}' column in Provider sheet.")
    except ValueError:
        print(f"'{col_name}' column not found, skipping dropdown validation.")

# Add dropdown validation for 'Board Certification 1' through 'Board Certification 5' columns in Provider sheet.
for i in range(1, 6):
    col_name = f'Board Certification {i}'
    try:
        board_col = header.index(col_name) + 1
        board_col_letter = get_column_letter(board_col)
        dv_board = DataValidation(type="list", formula1='=ValidationAndReference!$N$2:$N$299', allow_blank=True)
        dv_board_range = f"{board_col_letter}2:{board_col_letter}{ws.max_row}"
        dv_board.add(dv_board_range)
        ws.add_data_validation(dv_board)
        wb.save(merged_file_path)
        print(f"Added dropdown validation for '{col_name}' column in Provider sheet.")
    except ValueError:
        print(f"'{col_name}' column not found, skipping dropdown validation.")

# Add dropdown validation for 'Sub Board Certification 1' through 'Sub Board Certification 5' columns in Provider sheet.
for i in range(1, 6):
    col_name = f'Sub Board Certification {i}'
    try:
        sub_col = header.index(col_name) + 1
        sub_col_letter = get_column_letter(sub_col)
        dv_sub = DataValidation(type="list", formula1='=ValidationAndReference!$AB$2:$AB$156', allow_blank=True)
        dv_sub_range = f"{sub_col_letter}2:{sub_col_letter}{ws.max_row}"
        dv_sub.add(dv_sub_range)
        ws.add_data_validation(dv_sub)
        wb.save(merged_file_path)
        print(f"Added dropdown validation for '{col_name}' column in Provider sheet.")
    except ValueError:
        print(f"'{col_name}' column not found, skipping dropdown validation.")

# Add dropdown validation for 'Additional Languages Spoken 1' through 'Additional Languages Spoken 3' columns in Provider sheet.
for i in range(1, 4):
    col_name = f'Additional Languages Spoken {i}'
    try:
        lang_col = header.index(col_name) + 1
        lang_col_letter = get_column_letter(lang_col)
        dv_lang = DataValidation(type="list", formula1='=ValidationAndReference!$W$2:$W$144', allow_blank=True)
        dv_lang_range = f"{lang_col_letter}2:{lang_col_letter}{ws.max_row}"
        dv_lang.add(dv_lang_range)
        ws.add_data_validation(dv_lang)
        wb.save(merged_file_path)
        print(f"Added dropdown validation for '{col_name}' column in Provider sheet.")
    except ValueError:
        print(f"'{col_name}' column not found, skipping dropdown validation.")

# Add dropdown validation for 'Provider Type' column in Provider sheet.
try:
    provider_type_col = header.index('Provider Type') + 1
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    dv_provider_type = DataValidation(type="list", formula1='=ValidationAndReference!$Q$2:$Q$9', allow_blank=True)
    provider_type_col_letter = get_column_letter(provider_type_col)
    dv_provider_type_range = f"{provider_type_col_letter}2:{provider_type_col_letter}{ws.max_row}"
    dv_provider_type.add(dv_provider_type_range)
    ws.add_data_validation(dv_provider_type)
    # Set default value for all rows
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=provider_type_col, value='Practitioner - Full Profile')
    wb.save(merged_file_path)
    print("Added dropdown validation and default value for 'Provider Type' column in Provider sheet.")
except ValueError:
    print("'Provider Type' column not found, skipping dropdown validation.")

# Add dropdown validation for 'Enterprise Scheduling Flag' column in Provider sheet.
try:
    esf_col = header.index('Enterprise Scheduling Flag') + 1
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    dv_esf = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    esf_col_letter = get_column_letter(esf_col)
    dv_esf_range = f"{esf_col_letter}2:{esf_col_letter}{ws.max_row}"
    dv_esf.add(dv_esf_range)
    ws.add_data_validation(dv_esf)
    # Set default value for all rows
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=esf_col, value='No')
    wb.save(merged_file_path)
    print("Added dropdown validation and default value for 'Enterprise Scheduling Flag' column in Provider sheet.")
except ValueError:
    print("'Enterprise Scheduling Flag' column not found, skipping dropdown validation.")

# Add formula for 'Provider Type (Substatus) ID' column in Provider sheet.
try:
    substatus_col = header.index('Provider Type (Substatus) ID') + 1
    for row in range(2, ws.max_row + 1):
        formula = f'=IFERROR(INDEX(ValidationAndReference!P:P, MATCH(BE{row}, ValidationAndReference!Q:Q, 0)), "")'
        ws.cell(row=row, column=substatus_col, value=formula)
    wb.save(merged_file_path)
    print("Added formula for 'Provider Type (Substatus) ID' column in Provider sheet.")
except ValueError:
    print("'Provider Type (Substatus) ID' column not found, skipping formula step.")

# Add formulas for 'Professional Suffix ID 1', 'Professional Suffix ID 2', and 'Professional Suffix ID 3' columns in Provider sheet.
try:
    suffix_id_1_col = header.index('Professional Suffix ID 1') + 1
    suffix_id_2_col = header.index('Professional Suffix ID 2') + 1
    suffix_id_3_col = header.index('Professional Suffix ID 3') + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=suffix_id_1_col, value=f'=IFERROR(INDEX(ValidationAndReference!F:F, MATCH(D{row}, ValidationAndReference!G:G, 0)), "")')
        ws.cell(row=row, column=suffix_id_2_col, value=f'=IFERROR(INDEX(ValidationAndReference!F:F, MATCH(E{row}, ValidationAndReference!G:G, 0)), "")')
        ws.cell(row=row, column=suffix_id_3_col, value=f'=IFERROR(INDEX(ValidationAndReference!F:F, MATCH(F{row}, ValidationAndReference!G:G, 0)), "")')
    wb.save(merged_file_path)
    print("Added formulas for 'Professional Suffix ID 1/2/3' columns in Provider sheet.")
except ValueError:
    print("One or more 'Professional Suffix ID' columns not found, skipping formula step.")

# Add dropdown validation for 'Hospital Affiliation 1' through 'Hospital Affiliation 5' columns in Provider sheet.
for i in range(1, 6):
    col_name = f'Hospital Affiliation {i}'
    try:
        hosp_col = header.index(col_name) + 1
        hosp_col_letter = get_column_letter(hosp_col)
        dv_hosp = DataValidation(type="list", formula1='=ValidationAndReference!$T$2:$T$7258', allow_blank=True)
        dv_hosp_range = f"{hosp_col_letter}2:{hosp_col_letter}{ws.max_row}"
        dv_hosp.add(dv_hosp_range)
        ws.add_data_validation(dv_hosp)
        wb.save(merged_file_path)
        print(f"Added dropdown validation for '{col_name}' column in Provider sheet.")
    except ValueError:
        print(f"'{col_name}' column not found, skipping dropdown validation.")

# Add formulas for 'Language ID 1', 'Language ID 2', and 'Language ID 3' columns in Provider sheet.
try:
    lang_id_1_col = header.index('Language ID 1') + 1
    lang_id_2_col = header.index('Language ID 2') + 1
    lang_id_3_col = header.index('Language ID 3') + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=lang_id_1_col, value=f'=IFERROR(INDEX(ValidationAndReference!V:V, MATCH(AZ{row}, ValidationAndReference!W:W, 0)), "")')
        ws.cell(row=row, column=lang_id_2_col, value=f'=IFERROR(INDEX(ValidationAndReference!V:V, MATCH(BA{row}, ValidationAndReference!W:W, 0)), "")')
        ws.cell(row=row, column=lang_id_3_col, value=f'=IFERROR(INDEX(ValidationAndReference!V:V, MATCH(BB{row}, ValidationAndReference!W:W, 0)), "")')
    wb.save(merged_file_path)
    print("Added formulas for 'Language ID 1/2/3' columns in Provider sheet.")
except ValueError:
    print("One or more 'Language ID' columns not found, skipping formula step.")

# Step 5: Open Mergedoutput.xlsx automatically (Windows only)
# os.startfile(merged_file_path) # This line is moved to after suffix_check.py

# Add dropdown validations to Location sheet
try:
    wb_loc = openpyxl.load_workbook(merged_file_path)
    ws_loc = wb_loc["Location"]
    loc_header = [cell.value for cell in ws_loc[1]]
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    # Location Type dropdown
    try:
        loc_type_col = loc_header.index('Location Type') + 1
        loc_type_col_letter = get_column_letter(loc_type_col)
        dv_loc_type = DataValidation(type="list", formula1='"Virtual,In Person"', allow_blank=True)
        dv_loc_type_range = f"{loc_type_col_letter}2:{loc_type_col_letter}{ws_loc.max_row}"
        dv_loc_type.add(dv_loc_type_range)
        ws_loc.add_data_validation(dv_loc_type)
    except ValueError:
        print("'Location Type' column not found in Location sheet.")
    # State dropdown
    try:
        state_col = loc_header.index('State') + 1
        state_col_letter = get_column_letter(state_col)
        dv_state = DataValidation(type="list", formula1='=ValidationAndReference!$A$2:$A$55', allow_blank=True)
        dv_state_range = f"{state_col_letter}2:{state_col_letter}{ws_loc.max_row}"
        dv_state.add(dv_state_range)
        ws_loc.add_data_validation(dv_state)
    except ValueError:
        print("'State' column not found in Location sheet.")
    # Scheduling Software dropdown
    try:
        sched_col = loc_header.index('Scheduling Software') + 1
        sched_col_letter = get_column_letter(sched_col)
        dv_sched = DataValidation(type="list", formula1='=ValidationAndReference!$D$2:$D$750', allow_blank=True)
        dv_sched_range = f"{sched_col_letter}2:{sched_col_letter}{ws_loc.max_row}"
        dv_sched.add(dv_sched_range)
        ws_loc.add_data_validation(dv_sched)
    except ValueError:
        print("'Scheduling Software' column not found in Location sheet.")
    # Virtual Visit Type dropdown
    try:
        vvt_col = loc_header.index('Virtual Visit Type') + 1
        vvt_col_letter = get_column_letter(vvt_col)
        dv_vvt = DataValidation(type="list", formula1='=ValidationAndReference!$Y$2:$Y$3', allow_blank=True)
        dv_vvt_range = f"{vvt_col_letter}2:{vvt_col_letter}{ws_loc.max_row}"
        dv_vvt.add(dv_vvt_range)
        ws_loc.add_data_validation(dv_vvt)
    except ValueError:
        print("'Virtual Visit Type' column not found in Location sheet.")
    wb_loc.save(merged_file_path)
    print("Added dropdown validations to Location sheet.")
except Exception as e:
    print(f"Error adding dropdown validations to Location sheet: {e}")

# Now run suffix_check.py as the very last step
print("Running suffix_check.py to highlight invalid professional suffixes...")
subprocess.run(["python", "suffix_check.py"], check=True)
print("Finished highlighting invalid professional suffixes in Provider sheet.")

# --- Manual edit: Ensure all ZIP Codes in Location sheet are 5 digits (pad 4-digit with leading zero) ---
wb_loc = openpyxl.load_workbook(merged_file_path)
ws_loc = wb_loc["Location"]
header_row = [cell.value for cell in ws_loc[1]]
try:
    zip_col_idx = header_row.index('ZIP Code') + 1  # 1-based
    for row in range(2, ws_loc.max_row + 1):
        cell = ws_loc.cell(row=row, column=zip_col_idx)
        val = str(cell.value).strip() if cell.value is not None else ''
        if val.isdigit() and len(val) == 4:
            cell.value = f'0{val}'
    wb_loc.save(merged_file_path)
    print("Corrected 4-digit ZIP Codes in Location sheet to 5 digits.")
except ValueError:
    print("'ZIP Code' column not found in Location sheet, skipping ZIP correction.")

# --- Delete specified columns from Provider sheet ---
# Removed column deletion as requested.
#wb = openpyxl.load_workbook(merged_file_path)
#ws = wb['Provider']
#header = [cell.value for cell in ws[1]]
#columns_to_delete = [
#    'Facility Address', 'Facility City', 'Facility Zip', 'Facility State', 'Address line 2', 'Matched'
#]
#for col in columns_to_delete:
#    try:
#        idx = header.index(col) + 1
#        ws.delete_cols(idx)
#        header.pop(idx-1)
#    except ValueError:
#        pass
#wb.save(merged_file_path)

# Now open the file in Excel (Windows only)


wb = openpyxl.load_workbook(merged_file_path)
ws_provider = wb['Provider']
ws_location = wb['Location']
prov_header = [cell.value for cell in ws_provider[1]]
loc_header = [cell.value for cell in ws_location[1]]

loc_address_col = loc_header.index('Address line 1') + 1
loc_type_col = loc_header.index('Location Type') + 1
loc_cloud_id_col = loc_header.index('Location Cloud ID') + 1

try:
    provider_fac_addr_col = prov_header.index('Facility Address') + 1
    provider_locid1_col = prov_header.index('Location ID 1') + 1
    provider_locid2_col = prov_header.index('Location ID 2') + 1
except ValueError:
    provider_fac_addr_col = provider_locid1_col = provider_locid2_col = None

if None not in (provider_locid1_col, provider_locid2_col):
    for row in range(2, ws_provider.max_row + 1):
        locid1 = ws_provider.cell(row=row, column=provider_locid1_col).value
        if not locid1 or str(locid1).strip() == '':
            facility_addr = ws_provider.cell(row=row, column=provider_fac_addr_col).value
            loc_address_list = [ws_location.cell(row=lrow, column=loc_address_col).value for lrow in range(2, ws_location.max_row + 1)]
            fuzzy_matches = process.extract(facility_addr, loc_address_list, scorer=fuzz.token_sort_ratio, score_cutoff=60)
            best_inperson_id = None
            best_virtual_id = None
            for match_addr, score, lrow_offset in fuzzy_matches:
                lrow = lrow_offset + 2
                loc_type = ws_location.cell(row=lrow, column=loc_type_col).value
                loc_cloud_id = ws_location.cell(row=lrow, column=loc_cloud_id_col).value
                if loc_type == 'In Person' and not best_inperson_id:
                    best_inperson_id = loc_cloud_id
                if loc_type == 'Virtual' and not best_virtual_id:
                    best_virtual_id = loc_cloud_id
            if best_inperson_id:
                ws_provider.cell(row=row, column=provider_locid1_col, value=best_inperson_id)
            if best_virtual_id:
                ws_provider.cell(row=row, column=provider_locid2_col, value=best_virtual_id)
wb.save(merged_file_path)
print("Filled missing Location ID 1/2 in Provider sheet using fuzzy Facility Address mapping to Location sheet.")

wb = openpyxl.load_workbook(merged_file_path)
ws_provider = wb['Provider']
prov_header = [cell.value for cell in ws_provider[1]]
try:
    provider_locid1_col = prov_header.index('Location ID 1') + 1
    provider_locid2_col = prov_header.index('Location ID 2') + 1
except ValueError:
    provider_locid1_col = provider_locid2_col = None

if None not in (provider_locid1_col, provider_locid2_col):
    for row in range(2, ws_provider.max_row + 1):
        id1 = ws_provider.cell(row=row, column=provider_locid1_col).value
        id2 = ws_provider.cell(row=row, column=provider_locid2_col).value
        if (not id1 or str(id1).strip() == "") and id2 and str(id2).strip() != "":
            ws_provider.cell(row=row, column=provider_locid1_col, value=id2)
            ws_provider.cell(row=row, column=provider_locid2_col, value=None)
wb.save(merged_file_path)
print("Shifted Location ID 2 to Location ID 1 when Location ID 1 was missing.")

# Highlight Professional Statement cells over 2000 chars or containing URLs
wb = openpyxl.load_workbook(merged_file_path)
ws_provider = wb['Provider']
prov_header = [cell.value for cell in ws_provider[1]]
try:
    prof_stmt_col = prov_header.index('Professional Statement') + 1
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    url_pattern = re.compile(r'https?://|www\\.')
    for row in range(2, ws_provider.max_row + 1):
        cell = ws_provider.cell(row=row, column=prof_stmt_col)
        val = str(cell.value) if cell.value is not None else ''
        if len(val) > 2000 or url_pattern.search(val):
            cell.fill = yellow_fill
    wb.save(merged_file_path)
    print("Highlighted 'Professional Statement' cells >2000 chars or containing URLs in Provider sheet.")
except ValueError:
    print("'Professional Statement' column not found, skipping highlighting step.")

wb = openpyxl.load_workbook(merged_file_path)
ws_provider = wb['Provider']
ws_location = wb['Location']
prov_header = [cell.value for cell in ws_provider[1]]
loc_header = [cell.value for cell in ws_location[1]]

try:
    provider_locid1_col = prov_header.index('Location ID 1') + 1
    provider_practice_name_col = prov_header.index('Practice Name') + 1
    loc_cloud_id_col = loc_header.index('Location Cloud ID') + 1
    loc_practice_name_col = loc_header.index('Practice Name') + 1
except ValueError:
    provider_locid1_col = provider_practice_name_col = loc_cloud_id_col = loc_practice_name_col = None

if None not in (provider_locid1_col, provider_practice_name_col, loc_cloud_id_col, loc_practice_name_col):
    for row in range(2, ws_provider.max_row + 1):
        locid1 = ws_provider.cell(row=row, column=provider_locid1_col).value
        if locid1 and str(locid1).strip() != '':
            for lrow in range(2, ws_location.max_row + 1):
                loc_cloud_id = ws_location.cell(row=lrow, column=loc_cloud_id_col).value
                if locid1 == loc_cloud_id:
                    practice_name = ws_location.cell(row=lrow, column=loc_practice_name_col).value
                    ws_provider.cell(row=row, column=provider_practice_name_col, value=practice_name)
                    break
wb.save(merged_file_path)
print("Brought 'Practice Name' from Location sheet to Provider sheet after Location ID mapping.")

# --- FILLING 'Practice Cloud ID' in the Provider sheet from Location sheet ---
wb = openpyxl.load_workbook(merged_file_path)
ws_provider = wb['Provider']
ws_location = wb['Location']
prov_header = [cell.value for cell in ws_provider[1]]
loc_header = [cell.value for cell in ws_location[1]]
try:
    provider_locid1_col = prov_header.index('Location ID 1') + 1
    provider_practice_cloud_id_col = prov_header.index('Practice Cloud ID') + 1
    loc_cloud_id_col = loc_header.index('Location Cloud ID') + 1
    loc_practice_cloud_id_col = loc_header.index('Practice Cloud ID') + 1
except ValueError:
    provider_locid1_col = provider_practice_cloud_id_col = loc_cloud_id_col = loc_practice_cloud_id_col = None

if None not in (provider_locid1_col, provider_practice_cloud_id_col, loc_cloud_id_col, loc_practice_cloud_id_col):
    for row in range(2, ws_provider.max_row + 1):
        locid1 = ws_provider.cell(row=row, column=provider_locid1_col).value
        cloud_id = ''
        if locid1 and str(locid1).strip() != '':
            for lrow in range(2, ws_location.max_row + 1):
                loc_cloud_id = ws_location.cell(row=lrow, column=loc_cloud_id_col).value
                if locid1 == loc_cloud_id:
                    cloud_id = ws_location.cell(row=lrow, column=loc_practice_cloud_id_col).value
                    break
        ws_provider.cell(row=row, column=provider_practice_cloud_id_col, value=cloud_id)
    wb.save(merged_file_path)
    print("Filled 'Practice Cloud ID' in Provider sheet from Location sheet.")


def highlight_duplicate_npi(merged_file_path):
    """
    Highlights duplicate entries in the 'NPI Number' column of the Provider sheet in blue (#9BD7FF).
    """
    from openpyxl.styles import PatternFill
    wb = openpyxl.load_workbook(merged_file_path)
    ws = wb["Provider"]
    header = [cell.value for cell in ws[1]]
    try:
        npi_col = header.index("NPI Number") + 1
    except ValueError:
        print("'NPI Number' column not found in Provider sheet for duplicate highlighting.")
        return
    npi_count = {}
    # Count each NPI
    for row in range(2, ws.max_row + 1):
        npi = ws.cell(row=row, column=npi_col).value
        if npi is not None and str(npi).strip() != "":
            npi_count[npi] = npi_count.get(npi, 0) + 1
    # Blue fill for duplicates
    blue_fill = PatternFill(start_color='9BD7FF', end_color='9BD7FF', fill_type='solid')
    for row in range(2, ws.max_row + 1):
        npi = ws.cell(row=row, column=npi_col).value
        if npi is not None and npi_count.get(npi, 0) > 1:
            ws.cell(row=row, column=npi_col).fill = blue_fill
    wb.save(merged_file_path)
    print("Highlighted duplicate NPI Numbers in Provider sheet with #9BD7FF.")

# === Call highlight_duplicate_npi after all Provider sheet operations, before final print/statements ===
highlight_duplicate_npi(merged_file_path)



print("Running Location_2.py for post-processing...")
subprocess.run(["python", "Location_2.py"], check=True)
print("Location_2.py completed.")


