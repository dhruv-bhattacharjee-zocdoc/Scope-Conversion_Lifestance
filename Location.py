import openpyxl
import os
import subprocess
import re
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# --- Add smart_title_case helper ---
def smart_title_case(text, keep_upper=None):
    if not isinstance(text, str):
        return text
    if keep_upper is None:
        keep_upper = {'NE', 'SE', 'SW', 'NW', 'N', 'S', 'E', 'W'}
    # Always keep 2-letter state abbreviations uppercase
    # (Assume state abbreviations are always 2 letters and in keep_upper)
    def fix_word(word):
        w = word.strip()
        if w.upper() in keep_upper or (len(w) == 2 and w.isupper()):
            return w.upper()
        return w[:1].upper() + w[1:].lower() if w else w
    # Split on space and also handle hyphens
    def fix_token(token):
        return '-'.join(fix_word(part) for part in token.split('-'))
    return ' '.join(fix_token(token) for token in text.split())

# Define file paths
TEMPLATE_FILE = os.path.join(
    'Excel Files',
    'New Business Scope Sheet - Practice Locations and Providers.xlsx'
)
OUTPUT_FILE = os.path.join(
    'Excel Files',
    'Location.xlsx'
)
INPUT_FILE = os.path.join(
    'Excel Files',
    'Input.xlsx'
)

# Load the template workbook
wb_template = openpyxl.load_workbook(TEMPLATE_FILE, data_only=True)

# Create a new workbook for output
wb_output = openpyxl.Workbook()
# Remove the default sheet
if 'Sheet' in wb_output.sheetnames:
    std = wb_output['Sheet']
    wb_output.remove(std)

# List of sheets to copy
sheets_to_copy = ['Location', 'ValidationAndReference']

for sheet_name in sheets_to_copy:
    if sheet_name in wb_template.sheetnames:
        ws_template = wb_template[sheet_name]
        ws_output = wb_output.create_sheet(title=sheet_name)
        for row in ws_template.iter_rows(values_only=False):
            ws_output.append([cell.value for cell in row])
        # Copy column widths
        for col_letter, col_dim in ws_template.column_dimensions.items():
            ws_output.column_dimensions[col_letter].width = col_dim.width
        # Copy row heights
        for row_num, row_dim in ws_template.row_dimensions.items():
            ws_output.row_dimensions[row_num].height = row_dim.height
    else:
        print(f"Sheet '{sheet_name}' not found in template file.")

# --- Copy 'Facility Zip' from input to 'Zip Code' in output ---
# Load input workbook and get 'Facility Zip' column
wb_input = openpyxl.load_workbook(INPUT_FILE, data_only=True)
ws_input = wb_input.active
if ws_input is None:
    raise ValueError("Input worksheet could not be loaded.")
input_header_row = next(ws_input.iter_rows(min_row=1, max_row=1, values_only=True), [])
input_header = list(input_header_row) if input_header_row else []
try:
    facility_zip_idx = input_header.index('Facility Zip')
except ValueError:
    raise ValueError("'Facility Zip' column not found in input file.")
facility_zip_values = [row[facility_zip_idx] for row in ws_input.iter_rows(min_row=2, values_only=True)]

# Load output workbook and get 'Zip Code' column in 'Location' sheet
ws_location = wb_output['Location']
if ws_location is None:
    raise ValueError("'Location' sheet could not be loaded from output workbook.")
location_header_row = next(ws_location.iter_rows(min_row=1, max_row=1, values_only=True), [])
location_header = list(location_header_row) if location_header_row else []
try:
    zip_code_idx = location_header.index('ZIP Code')
except ValueError:
    raise ValueError("'ZIP Code' column not found in output file's Location sheet.")

# Write Facility Zip values to ZIP Code column in Location sheet
for i, value in enumerate(facility_zip_values, start=2):
    zip5 = str(value).split('-')[0] if value is not None else ''
    ws_location.cell(row=i, column=zip_code_idx+1, value=zip5)

def map_location_type(value):
    if value == 'Telehealth':
        return 'Virtual'
    elif value == 'In-Office':
        return 'In Person'
    elif value == 'Both':
        return 'Both'
    return value

# Map of input column to output column
column_mappings = [
    ('Facility Zip', 'ZIP Code', lambda v: str(v).split('-')[0] if v is not None else ''),
    ('Facility Address', 'Address line 1', lambda v: v),
    ('Facility City', 'City', lambda v: v),
    ('Facility State', 'State', lambda v: v),
    ('Telehealth or In-Office or Both', 'Location Type', map_location_type),
]

# For each mapping, copy values from input to output
for input_col, output_col, transform in column_mappings:
    try:
        input_idx = input_header.index(input_col)
    except ValueError:
        raise ValueError(f"'{input_col}' column not found in input file.")
    try:
        output_idx = location_header.index(output_col)
    except ValueError:
        raise ValueError(f"'{output_col}' column not found in output file's Location sheet.")
    values = [row[input_idx] for row in ws_input.iter_rows(min_row=2, values_only=True)]
    for i, value in enumerate(values, start=2):
        ws_location.cell(row=i, column=output_idx+1, value=transform(value))

# --- Address Standardization and Cleaning ---
# Load suffix mapping from reference file
SUFFIX_FILE = os.path.join('Excel Files', 'C1 Street Suffix Abbreviations.xlsx')
suffix_wb = openpyxl.load_workbook(SUFFIX_FILE, data_only=True)
suffix_ws = suffix_wb['Sheet1']
suffix_map = {}
for row in suffix_ws.iter_rows(min_row=2, values_only=True):
    common, usps = row[0], row[1]
    if common and usps:
        suffix_map[str(common).strip().lower()] = str(usps).strip()

# Helper regex for suite/unit/PO Box info
suite_keywords = [
    r"suite", r"ste", r"apt", r"apartment", r"floor", r"fl", r"unit", r"room", r"rm", r"bldg", r"#", r"p\.o\. box", r"po box"
]
suite_pattern = re.compile(r"(" + r"|".join(suite_keywords) + r").*", re.IGNORECASE)

# Define a fill for marking formatted cells
highlight_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

# Process addresses in wb_output['Location'] before saving
ws_location = wb_output['Location']
loc_header_row = next(ws_location.iter_rows(min_row=1, max_row=1, values_only=True), [])
loc_header = list(loc_header_row) if loc_header_row else []

try:
    addr1_idx = loc_header.index('Address line 1')
    addr2_idx = loc_header.index('Address line 2 (Office/Suite #)')
except ValueError as e:
    raise ValueError(f"Address column not found: {e}")

for i, row in enumerate(ws_location.iter_rows(min_row=2, max_row=ws_location.max_row), start=2):
    addr1_cell = row[addr1_idx]
    addr2_cell = row[addr2_idx]
    addr1_val = addr1_cell.value if addr1_cell else ''
    addr2_val = addr2_cell.value if addr2_cell else ''
    if not addr1_val:
        continue
    addr1_str = str(addr1_val)
    formatted = False
    # If there's a comma, move everything after the first comma to Address line 2
    if ',' in addr1_str:
        before_comma, after_comma = addr1_str.split(',', 1)
        addr1_str = before_comma.strip()
        after_comma = after_comma.strip()
        if after_comma:
            if addr2_val:
                addr2_str = str(addr2_val)
                new_addr2 = f"{addr2_str} {after_comma}".strip()
            else:
                new_addr2 = after_comma
            # Apply smart title case to Address line 2
            ws_location.cell(row=i, column=addr2_idx+1, value=smart_title_case(new_addr2))
    # Move suite/unit/PO Box info to Address line 2
    match = suite_pattern.search(addr1_str)
    if match:
        suite_part = addr1_str[match.start():].strip()
        addr1_str = addr1_str[:match.start()].strip()
        # Append to Address line 2, preserving existing content
        if addr2_val:
            addr2_str = str(addr2_val)
            new_addr2 = f"{addr2_str} {suite_part}".strip()
        else:
            new_addr2 = suite_part
        ws_location.cell(row=i, column=addr2_idx+1, value=smart_title_case(new_addr2))
    # Standardize street suffix (last word)
    words = addr1_str.split()
    if words:
        last_word = words[-1].rstrip('.')
        last_word_lower = last_word.lower()
        if last_word_lower in suffix_map:
            words[-1] = suffix_map[last_word_lower]
            addr1_str = ' '.join(words)
            formatted = True
    # Write cleaned Address line 1 (with smart title case)
    cell = ws_location.cell(row=i, column=addr1_idx+1, value=smart_title_case(addr1_str))
    if formatted:
        cell.fill = highlight_fill
    # Also apply smart title case to Address line 2 if not already set above
    if not (',' in str(addr1_val) or (suite_pattern.search(str(addr1_val)))):
        if addr2_val:
            ws_location.cell(row=i, column=addr2_idx+1, value=smart_title_case(str(addr2_val)))

print("Address fields standardized and cleaned in Location sheet.")

# After all address cleaning, create the Combined address column
try:
    combined_idx = loc_header.index('Combined address')
except ValueError:
    combined_idx = None
    # If not present, add it as the last column
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Combined address')
    combined_idx = ws_location.max_column - 1  # 0-based index

# Get indices for required columns
try:
    city_idx = loc_header.index('City')
    state_idx = loc_header.index('State')
    zip_idx = loc_header.index('ZIP Code')
except ValueError as e:
    raise ValueError(f"Required column not found: {e}")

for i, row in enumerate(ws_location.iter_rows(min_row=2, max_row=ws_location.max_row), start=2):
    addr1 = row[addr1_idx].value if row[addr1_idx] else ''
    city = row[city_idx].value if row[city_idx] else ''
    state = row[state_idx].value if row[state_idx] else ''
    zipcode = row[zip_idx].value if row[zip_idx] else ''
    # Only write combined address if at least one field is non-empty
    if any([addr1, city, state, zipcode]):
        addr1_cased = smart_title_case(addr1)
        city_cased = smart_title_case(city)
        state_cased = str(state).upper() if state else ''
        combined = f"{addr1_cased}, {city_cased}, {state_cased} {zipcode}".strip().replace('  ', ' ')
        ws_location.cell(row=i, column=combined_idx+1, value=combined)
    else:
        ws_location.cell(row=i, column=combined_idx+1, value='')

# Add or find the 'Show name in search?' column
try:
    show_name_idx = loc_header.index('Show name in search?')
except ValueError:
    show_name_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Show name in search?')
    show_name_idx = ws_location.max_column - 1  # 0-based index

# Find the 'Location Name' column (should be column A)
try:
    location_name_idx = loc_header.index('Location Name')
except ValueError as e:
    raise ValueError(f"'Location Name' column not found: {e}")

for i in range(2, ws_location.max_row+1):
    # Column A is 1-based index 1
    formula = f'=IF(A{i}<>"", "Yes", "")'
    ws_location.cell(row=i, column=show_name_idx+1, value=formula)

# Add dropdown for 'Show name in search?' column (Yes/No)
show_name_col_letter = get_column_letter(show_name_idx+1)
dv_show_name = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
dv_show_name_range = f"{show_name_col_letter}2:{show_name_col_letter}{ws_location.max_row}"
dv_show_name.add(dv_show_name_range)
ws_location.add_data_validation(dv_show_name)

# Add or find the 'Complete Location' column
try:
    complete_loc_idx = loc_header.index('Complete Location')
except ValueError:
    complete_loc_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Complete Location')
    complete_loc_idx = ws_location.max_column - 1  # 0-based index

for i in range(2, ws_location.max_row+1):
    # Formula: =IF(AND(A2<>"", B2<>"", F2<>"", I2<>""), A2, "")
    formula = f'=IF(AND(A{i}<>"", B{i}<>"", F{i}<>"", I{i}<>""), A{i}, "")'
    ws_location.cell(row=i, column=complete_loc_idx+1, value=formula)

# Add or find the 'Scheduling Software ID' column
try:
    sched_id_idx = loc_header.index('Scheduling Software ID')
except ValueError:
    sched_id_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Scheduling Software ID')
    sched_id_idx = ws_location.max_column - 1  # 0-based index

for i in range(2, ws_location.max_row+1):
    # Formula: =IF(ISBLANK(S2),"",INDEX(ValidationAndReference!C:C,MATCH(S2,ValidationAndReference!D:D,0)))
    formula = f'=IF(ISBLANK(S{i}),"",INDEX(ValidationAndReference!C:C,MATCH(S{i},ValidationAndReference!D:D,0)))'
    ws_location.cell(row=i, column=sched_id_idx+1, value=formula)

# Add dropdown for 'Virtual Visit Type' in Location sheet using values from ValidationAndReference
valref_ws = wb_output['ValidationAndReference']
valref_header_row = next(valref_ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
valref_header = list(valref_header_row) if valref_header_row else []
try:
    vvt_idx = valref_header.index('Virtual Visit Type')
except ValueError:
    vvt_idx = None
    raise ValueError("'Virtual Visit Type' column not found in ValidationAndReference sheet.")
# Collect unique, non-empty values (including all possible values in the column, even if repeated or with extra whitespace)
vvt_values = set()
for row in valref_ws.iter_rows(min_row=2, max_row=valref_ws.max_row, values_only=True):
    val = row[vvt_idx] if vvt_idx is not None and vvt_idx < len(row) else None
    if val is not None and str(val).strip() != '':
        vvt_values.add(str(val).strip())
vvt_list = sorted(vvt_values)

# Create a dynamic named range for Virtual Visit Type options in ValidationAndReference
vvt_col_letter = get_column_letter(vvt_idx+1)
# Find the last non-empty row in the Virtual Visit Type column
last_vvt_row = valref_ws.max_row
for r in range(valref_ws.max_row, 1, -1):
    cell_val = valref_ws.cell(row=r, column=vvt_idx+1).value
    if cell_val is not None and str(cell_val).strip() != '':
        last_vvt_row = r
        break
# Define the named range (excluding header)
vvt_range = f"ValidationAndReference!${vvt_col_letter}$2:${vvt_col_letter}${last_vvt_row}"
wb_output.defined_names.add(DefinedName('VirtualVisitTypeList', attr_text=vvt_range))

# Find or add the 'Virtual Visit Type' column in Location sheet
try:
    loc_vvt_idx = loc_header.index('Virtual Visit Type')
except ValueError:
    loc_vvt_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Virtual Visit Type')
    loc_vvt_idx = ws_location.max_column - 1  # 0-based index

# Set up data validation dropdown using the named range
col_letter = get_column_letter(loc_vvt_idx+1)
dv = DataValidation(type="list", formula1='=VirtualVisitTypeList', allow_blank=True)
dv_range = f"{col_letter}2:{col_letter}{ws_location.max_row}"
dv.add(dv_range)
ws_location.add_data_validation(dv)

# Add dropdown for 'State' in Location sheet using values from ValidationAndReference 'State Lookup'
try:
    state_lookup_idx = valref_header.index('State Lookup')
except ValueError:
    state_lookup_idx = None
    raise ValueError("'State Lookup' column not found in ValidationAndReference sheet.")

# Find the last non-empty row in the State Lookup column
state_col_letter = get_column_letter(state_lookup_idx+1)
last_state_row = valref_ws.max_row
for r in range(valref_ws.max_row, 1, -1):
    cell_val = valref_ws.cell(row=r, column=state_lookup_idx+1).value
    if cell_val is not None and str(cell_val).strip() != '':
        last_state_row = r
        break
# Define the named range (excluding header)
state_range = f"ValidationAndReference!${state_col_letter}$2:${state_col_letter}${last_state_row}"
wb_output.defined_names.add(DefinedName('StateLookupList', attr_text=state_range))

# Find the 'State' column in Location sheet
try:
    loc_state_idx = loc_header.index('State')
except ValueError:
    loc_state_idx = None
    raise ValueError("'State' column not found in Location sheet.")

# Set up data validation dropdown using the named range
state_col_letter_loc = get_column_letter(loc_state_idx+1)
dv_state = DataValidation(type="list", formula1='=StateLookupList', allow_blank=True)
dv_state_range = f"{state_col_letter_loc}2:{state_col_letter_loc}{ws_location.max_row}"
dv_state.add(dv_state_range)
ws_location.add_data_validation(dv_state)

# Add dropdown for 'Scheduling Software' in Location sheet using values from ValidationAndReference 'Software List'
try:
    software_list_idx = valref_header.index('Software List')
except ValueError:
    software_list_idx = None
    raise ValueError("'Software List' column not found in ValidationAndReference sheet.")

# Find the last non-empty row in the Software List column
software_col_letter = get_column_letter(software_list_idx+1)
last_software_row = valref_ws.max_row
for r in range(valref_ws.max_row, 1, -1):
    cell_val = valref_ws.cell(row=r, column=software_list_idx+1).value
    if cell_val is not None and str(cell_val).strip() != '':
        last_software_row = r
        break
# Define the named range (excluding header)
software_range = f"ValidationAndReference!${software_col_letter}$2:${software_col_letter}${last_software_row}"
wb_output.defined_names.add(DefinedName('SoftwareList', attr_text=software_range))

# Find the 'Scheduling Software' column in Location sheet
try:
    loc_software_idx = loc_header.index('Scheduling Software')
except ValueError:
    loc_software_idx = None
    raise ValueError("'Scheduling Software' column not found in Location sheet.")

# Set up data validation dropdown using the named range
software_col_letter_loc = get_column_letter(loc_software_idx+1)
dv_software = DataValidation(type="list", formula1='=SoftwareList', allow_blank=True)
dv_software_range = f"{software_col_letter_loc}2:{software_col_letter_loc}{ws_location.max_row}"
dv_software.add(dv_software_range)
ws_location.add_data_validation(dv_software)

# Add dropdown for 'Practice Name' in Location sheet using values from ValidationAndReference!$AD$2:$AD$430
try:
    practice_name_idx = loc_header.index('Practice Name')
except ValueError:
    practice_name_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Practice Name')
    practice_name_idx = ws_location.max_column - 1  # 0-based index

practice_name_col_letter = get_column_letter(practice_name_idx+1)
dv_practice_name = DataValidation(type="list", formula1='=ValidationAndReference!$AD$2:$AD$430', allow_blank=True)
dv_practice_name_range = f"{practice_name_col_letter}2:{practice_name_col_letter}{ws_location.max_row}"
dv_practice_name.add(dv_practice_name_range)
ws_location.add_data_validation(dv_practice_name)

# Also apply smart title case to City column
for i, row in enumerate(ws_location.iter_rows(min_row=2, max_row=ws_location.max_row), start=2):
    city_cell = row[city_idx]
    if city_cell and city_cell.value:
        ws_location.cell(row=i, column=city_idx+1, value=smart_title_case(str(city_cell.value)))

# --- Duplicate rows for 'Both' in Location Type ---
loc_header_row = next(ws_location.iter_rows(min_row=1, max_row=1, values_only=True), [])
loc_header = list(loc_header_row) if loc_header_row else []
try:
    location_type_idx = loc_header.index('Location Type')
except ValueError:
    location_type_idx = None
    raise ValueError("'Location Type' column not found in Location sheet.")

rows_to_duplicate = []
for i, row in enumerate(ws_location.iter_rows(min_row=2, max_row=ws_location.max_row, values_only=True), start=2):
    if row[location_type_idx] == 'Both':
        rows_to_duplicate.append((i, row))

# To avoid index shifting, process from bottom up
for i, row in reversed(rows_to_duplicate):
    # Remove the original row
    ws_location.delete_rows(i)
    # Insert two new rows: one with 'Virtual', one with 'In Person'
    new_row_virtual = list(row)
    new_row_virtual[location_type_idx] = 'Virtual'
    new_row_inperson = list(row)
    new_row_inperson[location_type_idx] = 'In Person'
    ws_location.insert_rows(i)
    for col_idx, value in enumerate(new_row_inperson, start=1):
        ws_location.cell(row=i, column=col_idx, value=value)
    ws_location.insert_rows(i)
    for col_idx, value in enumerate(new_row_virtual, start=1):
        ws_location.cell(row=i, column=col_idx, value=value)
# --- End duplication logic ---

# Save the output workbook
wb_output.save(OUTPUT_FILE)
print(f"Sheets {sheets_to_copy} copied to {OUTPUT_FILE} and 'Facility Zip' copied to 'Zip Code'.")

# Open the output file automatically
#output_path_abs = os.path.abspath(OUTPUT_FILE)
#subprocess.Popen(['start', '', output_path_abs], shell=True)
