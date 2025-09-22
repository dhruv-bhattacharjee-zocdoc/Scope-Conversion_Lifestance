import openpyxl
import os
from Name import extract_name_gender, add_gender_dropdown
from Npi import extract_npi
from Headshot import extract_headshot
from professional_suffix import extract_professional_suffix
from Specialty import extract_specialty, add_specialty_dropdowns
from PatientsAccepted import extract_patients_accepted, set_patients_accepted_dropdown
from Education import extract_education
from Professional_statement import extract_professional_statement
from Board_certification import extract_board_certification, extract_board_subspecialty, set_board_certification_dropdown
from optoutrating import set_opt_out_of_ratings_dropdown
from ESF import set_enterprise_scheduling_flag_dropdown
from Langauge import extract_languages
import subprocess
from specialtydropdown import add_specialty_valref_dropdowns
from Extract_NPI import create_npi_specialty_excel
import re

# Define the input, template, and output file paths
input_file = r"Excel Files/Input.xlsx"
template_file = r"C:\Users\dhruv.bhattacharjee\Desktop\PDO Data Transposition\Scope Conversion_Lifestance\Excel Files\New Business Scope Sheet - Practice Locations and Providers.xlsx"
output_file = r"C:\Users\dhruv.bhattacharjee\Desktop\PDO Data Transposition\Scope Conversion_Lifestance\Excel Files\Output.xlsx"

# Extract name and gender data using Name.py
extracted_rows = extract_name_gender(input_file)
# Extract NPI data using Npi.py
npi_list = extract_npi(input_file)
# Extract Headshot URL data using Headshot.py
headshot_list = extract_headshot(input_file)
# Extract Professional Suffix data using professional_suffix.py
suffix_lists = extract_professional_suffix(input_file)
# Extract specialty data using specialty.py
specialty_list = extract_specialty(input_file)
# Extract Patients Accepted data using PatientsAccepted.py
patients_accepted_list = extract_patients_accepted(input_file)
# Extract Education data using Education.py
education_list = extract_education(input_file)
# Extract Professional Statement data using Professional_statement.py
professional_statement_list = extract_professional_statement(input_file)
# Extract Board Certification data using Board_certification.py
board_certification_list = extract_board_certification(input_file)
board_subspecialty_list = extract_board_subspecialty(input_file)
# Extract Languages data using Langauge.py
lang1_list, lang2_list = extract_languages(input_file)
# Extract Facility Address data from Input.xlsx
import openpyxl
# After extracting other lists
wb_in = openpyxl.load_workbook(input_file)
ws_in = wb_in.active
input_header_row = [cell.value for cell in ws_in[1]]
try:
    facility_address_idx = input_header_row.index('Facility Address')
except ValueError:
    facility_address_idx = None
facility_address_list = []
if facility_address_idx is not None:
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        facility_address_list.append(row[facility_address_idx])
else:
    facility_address_list = [None] * len(extracted_rows)

# Extract Facility City, Facility Zip, Facility State from Input.xlsx
try:
    facility_city_idx = input_header_row.index('Facility City')
except ValueError:
    facility_city_idx = None
facility_city_list = []
if facility_city_idx is not None:
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        facility_city_list.append(row[facility_city_idx])
else:
    facility_city_list = [None] * len(extracted_rows)

try:
    facility_zip_idx = input_header_row.index('Facility Zip')
except ValueError:
    facility_zip_idx = None
facility_zip_list = []
if facility_zip_idx is not None:
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        zip_val = row[facility_zip_idx]
        if zip_val is not None:
            zip_val = str(zip_val)[:5]
        facility_zip_list.append(zip_val)
else:
    facility_zip_list = [None] * len(extracted_rows)

try:
    facility_state_idx = input_header_row.index('Facility State')
except ValueError:
    facility_state_idx = None
facility_state_list = []
if facility_state_idx is not None:
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        facility_state_list.append(row[facility_state_idx])
else:
    facility_state_list = [None] * len(extracted_rows)

# Load the template workbook and Provider sheet
wb_template = openpyxl.load_workbook(template_file)
ws_template = wb_template['Provider']

# Create a new workbook for output and copy the template structure
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
if ws_out is None:
    ws_out = wb_out.create_sheet(title='Provider')
else:
    ws_out.title = 'Provider'

# Define the columns to always include
extra_columns = ['Facility Address', 'Facility City', 'Facility Zip', 'Facility State', 'Address line 2']
# Copy the template headers
header_row = [cell.value for cell in ws_template[1]]
for col in extra_columns:
    if col not in header_row:
        header_row.append(col)
if ws_out is not None:
    ws_out.append(header_row)

# Helper for smart camel case (preserve NW, SW, SE, NE)
def smart_camel_case(text, keep_upper=None):
    if not isinstance(text, str):
        return text
    if keep_upper is None:
        keep_upper = {'NE', 'SE', 'SW', 'NW', 'N', 'S', 'E', 'W'}
    def fix_word(word):
        w = word.strip()
        if w.upper() in keep_upper:
            return w.upper()
        return w[:1].upper() + w[1:].lower() if w else w
    def fix_token(token):
        return '-'.join(fix_word(part) for part in token.split('-'))
    return ' '.join(fix_token(token) for token in text.split())

# Suite/unit/PO Box regex (same as Location.py)
suite_keywords = [
    r"suite", r"ste", r"apt", r"apartment", r"floor", r"fl", r"unit", r"room", r"rm", r"bldg", r"#", r"p\.o\. box", r"po box"
]
suite_pattern = re.compile(r"(" + r"|".join(suite_keywords) + r").*", re.IGNORECASE)

def split_and_clean_address(address):
    address_line_1 = address if address else ""
    address_line_2 = ""
    if address_line_1:
        addr_str = str(address_line_1)
        # If there's a comma, move everything after the first comma to Address line 2
        if ',' in addr_str:
            before_comma, after_comma = addr_str.split(',', 1)
            address_line_1 = before_comma.strip()
            address_line_2 = after_comma.strip()
        # Move suite/unit/PO Box info to Address line 2
        match = suite_pattern.search(address_line_1)
        if match:
            suite_part = address_line_1[match.start():].strip()
            address_line_1 = address_line_1[:match.start()].strip()
            if address_line_2:
                address_line_2 = f"{address_line_2} {suite_part}".strip()
            else:
                address_line_2 = suite_part
        # Camel case both, preserving NW, SW, SE, NE
        address_line_1 = smart_camel_case(address_line_1)
        address_line_2 = smart_camel_case(address_line_2) if address_line_2 else ""
    return address_line_1, address_line_2

# For each extracted row, create a row matching the template structure
for idx, extracted in enumerate(extracted_rows):
    new_row = []
    # Prepare address line 1 and 2 from Facility Address using locationmapping.py
    facility_address = facility_address_list[idx] if idx < len(facility_address_list) else ""
    address_line_1, address_line_2 = split_and_clean_address(facility_address)
    for header in header_row:
        if header in ['First Name', 'Last Name', 'Gender']:
            value = extracted.get(header, "")
            new_row.append(value if value is not None else "")
        elif header == 'NPI Number':
            value = npi_list[idx] if idx < len(npi_list) else ""
            new_row.append(value if value is not None else "")
        elif header == 'Headshot Link':
            value = headshot_list[idx] if idx < len(headshot_list) else ""
            new_row.append(value if value is not None else "")
        elif isinstance(header, str) and header.startswith('Professional Suffix '):
            # Only map if header is exactly 'Professional Suffix 1', 'Professional Suffix 2', etc.
            try:
                suffix_num = int(header.split('Professional Suffix ')[1]) - 1
                if suffix_num >= 0:
                    suffix_values = suffix_lists[idx] if idx < len(suffix_lists) else [""]
                    value = suffix_values[suffix_num] if suffix_num < len(suffix_values) else ""
                else:
                    value = ""
            except (IndexError, ValueError):
                value = ""
            new_row.append(value if value is not None else "")
        elif header == 'Specialty 1':
            value = ""
            new_row.append(value)
        elif header == 'Patients Accepted':
            value = patients_accepted_list[idx] if idx < len(patients_accepted_list) else ""
            new_row.append(value if value is not None else "")
        elif header == 'Education 1':
            value = education_list[idx] if idx < len(education_list) else ""
            new_row.append(value if value is not None else "")
        elif header == 'Professional Statement':
            value = professional_statement_list[idx] if idx < len(professional_statement_list) else ""
            new_row.append(value if value is not None else "")
        elif header.startswith('Board Certification') and header.split()[-1] in {'1','2','3','4','5'}:
            value = ""
            new_row.append(value)
        elif header.startswith('Sub Board Certification') and header.split()[-1] in {'1','2','3','4','5'}:
            value = ""
            new_row.append(value)
        elif header == 'Additional Languages Spoken 1':
            value = lang1_list[idx] if idx < len(lang1_list) else ""
            new_row.append(value if value is not None else "")
        elif header == 'Additional Languages Spoken 2':
            value = lang2_list[idx] if idx < len(lang2_list) else ""
            new_row.append(value if value is not None else "")
        elif header == 'Facility Address':
            new_row.append(address_line_1 if address_line_1 is not None else "")
        elif header == 'Address line 2':
            new_row.append(address_line_2 if address_line_2 is not None else "")
        elif header == 'Facility City':
            value = facility_city_list[idx] if idx < len(facility_city_list) else ""
            if value:
                value = str(value).title()
            new_row.append(value if value is not None else "")
        elif header == 'Facility Zip':
            value = facility_zip_list[idx] if idx < len(facility_zip_list) else ""
            new_row.append(value if value is not None else "")
        elif header == 'Facility State':
            value = facility_state_list[idx] if idx < len(facility_state_list) else ""
            new_row.append(value if value is not None else "")
        else:
            new_row.append("")
    if ws_out is not None:
        ws_out.append(new_row)

# Add gender dropdown to Provider sheet
add_gender_dropdown(ws_out)
# Save the output workbook
wb_out.save(output_file)

# Add dropdown for Patients Accepted
set_patients_accepted_dropdown(output_file)

# Add dropdown for Board Certification 1
set_board_certification_dropdown(output_file)

# Now copy 'ValidationAndReference' and 'Location' sheets from template to output
from openpyxl import load_workbook

def copy_sheet(src_wb, dest_wb, sheet_name):
    src_ws = src_wb[sheet_name]
    # Create new sheet in destination workbook
    dest_ws = dest_wb.create_sheet(title=sheet_name)
    for row in src_ws.iter_rows():
        dest_ws.append([cell.value for cell in row])
    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[col_letter].width = dim.width
    # Copy row heights
    for row_num, dim in src_ws.row_dimensions.items():
        dest_ws.row_dimensions[row_num].height = dim.height

# Reopen output file to add sheets
wb_out = load_workbook(output_file)
wb_template = load_workbook(template_file)
# Remove all sheets except 'Provider' (which is the main data sheet just created)
for sheet in wb_out.sheetnames:
    if sheet != 'Provider':
        std = wb_out[sheet]
        wb_out.remove(std)
# Copy only Provider and ValidationAndReference in the correct order (no Location)
for sheet in ['Provider', 'ValidationAndReference']:
    if sheet == 'Provider':
        # Rename the existing sheet to 'Provider' if not already
        ws = None
        if 'Provider' in wb_out.sheetnames:
            ws = wb_out['Provider']
        else:
            ws = wb_out.active
        if ws is not None:
            ws.title = 'Provider'
    elif sheet in wb_template.sheetnames:
        # Remove if already exists (to avoid duplicates)
        if sheet in wb_out.sheetnames:
            wb_out.remove(wb_out[sheet])
        # Copy from template
        copy_sheet(wb_template, wb_out, sheet)
# Reorder sheets using move_sheet
sheet_order = ['Provider', 'ValidationAndReference']
for idx, sheet_name in enumerate(sheet_order):
    if sheet_name in wb_out.sheetnames:
        wb_out.move_sheet(wb_out[sheet_name], offset=idx - wb_out.sheetnames.index(sheet_name))
wb_out.save(output_file)

# Add specialty dropdowns to Provider sheet (after ValidationAndReference sheet is present)
# Remove or comment out the following lines:
# ws_valref = wb_out['ValidationAndReference']
# add_specialty_dropdowns(ws_out, ws_valref)

# Call Location.py to generate the Location sheet
subprocess.run(['python', 'Location.py'], check=True)

# After Location.xlsx is generated, copy the 'Location' sheet into the main output file
location_wb = openpyxl.load_workbook('Excel Files/Location.xlsx')
location_ws = location_wb['Location']
wb_out = openpyxl.load_workbook(output_file)
# Remove existing 'Location' sheet if present
if 'Location' in wb_out.sheetnames:
    std = wb_out['Location']
    wb_out.remove(std)
# Create new 'Location' sheet and copy contents
ws_location = wb_out.create_sheet(title='Location')
for row in location_ws.iter_rows():
    ws_location.append([cell.value for cell in row])
# Copy column widths
for col_letter, dim in location_ws.column_dimensions.items():
    ws_location.column_dimensions[col_letter].width = dim.width
# Copy row heights
for row_num, dim in location_ws.row_dimensions.items():
    ws_location.row_dimensions[row_num].height = dim.height

# Autofill 'Location 1' in Provider sheet with 'Location Name' from Location sheet
provider_ws = wb_out['Provider']
location_ws = wb_out['Location']
provider_header = [cell.value for cell in provider_ws[1]]
location_header = [cell.value for cell in location_ws[1]]
try:
    location1_idx = provider_header.index('Location 1')
    location_name_idx = location_header.index('Location Name')
except ValueError:
    location1_idx = None
    location_name_idx = None

if location1_idx is not None and location_name_idx is not None:
    max_row = min(provider_ws.max_row, location_ws.max_row)
    for row in range(2, max_row+1):
        location_name = location_ws.cell(row=row, column=location_name_idx+1).value
        provider_ws.cell(row=row, column=location1_idx+1, value=location_name)

# Add dropdown for 'Practice Name' in Location sheet using values from ValidationAndReference!$AD$2:$AD$430
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
location_header = [cell.value for cell in ws_location[1]]
try:
    practice_name_idx = location_header.index('Practice Name')
except ValueError:
    practice_name_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Practice Name')
    location_header = [cell.value for cell in ws_location[1]]
    practice_name_idx = location_header.index('Practice Name')
practice_name_col_letter = get_column_letter(practice_name_idx+1)
dv_practice_name = DataValidation(type="list", formula1='=ValidationAndReference!$AD$2:$AD$430', allow_blank=True)
dv_practice_name_range = f"{practice_name_col_letter}2:{practice_name_col_letter}{ws_location.max_row}"
dv_practice_name.add(dv_practice_name_range)
ws_location.add_data_validation(dv_practice_name)

# Add dropdown for 'Location Type' in Location sheet with options 'Virtual' and 'In Person'
try:
    location_type_idx = location_header.index('Location Type')
except ValueError:
    location_type_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Location Type')
    location_header = [cell.value for cell in ws_location[1]]
    location_type_idx = location_header.index('Location Type')
location_type_col_letter = get_column_letter(location_type_idx+1)
dv_location_type = DataValidation(type="list", formula1='"Virtual,In Person"', allow_blank=True)
dv_location_type_range = f"{location_type_col_letter}2:{location_type_col_letter}{ws_location.max_row}"
dv_location_type.add(dv_location_type_range)
ws_location.add_data_validation(dv_location_type)

# Add dropdown for 'State' in Location sheet with source =ValidationAndReference!$A$2:$A$55
try:
    state_idx = location_header.index('State')
except ValueError:
    state_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='State')
    location_header = [cell.value for cell in ws_location[1]]
    state_idx = location_header.index('State')
state_col_letter = get_column_letter(state_idx+1)
dv_state = DataValidation(type="list", formula1='=ValidationAndReference!$A$2:$A$55', allow_blank=True)
dv_state_range = f"{state_col_letter}2:{state_col_letter}{ws_location.max_row}"
dv_state.add(dv_state_range)
ws_location.add_data_validation(dv_state)

# Add dropdown for 'Virtual Visit Type' in Location sheet with source =ValidationAndReference!$Y$2:$Y$3
try:
    vvt_idx = location_header.index('Virtual Visit Type')
except ValueError:
    vvt_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Virtual Visit Type')
    location_header = [cell.value for cell in ws_location[1]]
    vvt_idx = location_header.index('Virtual Visit Type')
vvt_col_letter = get_column_letter(vvt_idx+1)
dv_vvt = DataValidation(type="list", formula1='=ValidationAndReference!$Y$2:$Y$3', allow_blank=True)
dv_vvt_range = f"{vvt_col_letter}2:{vvt_col_letter}{ws_location.max_row}"
dv_vvt.add(dv_vvt_range)
ws_location.add_data_validation(dv_vvt)

# Add dropdown for 'Show name in search?' in Location sheet with options 'Yes' and 'No'
try:
    show_name_idx = location_header.index('Show name in search?')
except ValueError:
    show_name_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Show name in search?')
    location_header = [cell.value for cell in ws_location[1]]
    show_name_idx = location_header.index('Show name in search?')
show_name_col_letter = get_column_letter(show_name_idx+1)
dv_show_name = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
dv_show_name_range = f"{show_name_col_letter}2:{show_name_col_letter}{ws_location.max_row}"
dv_show_name.add(dv_show_name_range)
ws_location.add_data_validation(dv_show_name)

# Add dropdown for 'Scheduling Software' in Location sheet with source =ValidationAndReference!$D$2:$D$750
try:
    sched_software_idx = location_header.index('Scheduling Software')
except ValueError:
    sched_software_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Scheduling Software')
    location_header = [cell.value for cell in ws_location[1]]
    sched_software_idx = location_header.index('Scheduling Software')
sched_software_col_letter = get_column_letter(sched_software_idx+1)
dv_sched_software = DataValidation(type="list", formula1='=ValidationAndReference!$D$2:$D$750', allow_blank=True)
dv_sched_software_range = f"{sched_software_col_letter}2:{sched_software_col_letter}{ws_location.max_row}"
dv_sched_software.add(dv_sched_software_range)
ws_location.add_data_validation(dv_sched_software)

# Add formula for 'Complete Location' in Location sheet
try:
    complete_loc_idx = location_header.index('Complete Location')
except ValueError:
    complete_loc_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Complete Location')
    location_header = [cell.value for cell in ws_location[1]]
    complete_loc_idx = location_header.index('Complete Location')
for i in range(2, ws_location.max_row+1):
    formula = f'=IF(AND(B{i}<>"", C{i}<>"", G{i}<>"", J{i}<>""), A{i}&" - "&B{i}, "")'
    ws_location.cell(row=i, column=complete_loc_idx+1, value=formula)

# Add formula for 'Scheduling Software ID' in Location sheet
try:
    sched_id_idx = location_header.index('Scheduling Software ID')
except ValueError:
    sched_id_idx = None
    ws_location.cell(row=1, column=ws_location.max_column+1, value='Scheduling Software ID')
    location_header = [cell.value for cell in ws_location[1]]
    sched_id_idx = location_header.index('Scheduling Software ID')
for i in range(2, ws_location.max_row+1):
    formula = f'=IF(ISBLANK(U{i}),"",INDEX(ValidationAndReference!C:C,MATCH(U{i},ValidationAndReference!D:D,0)))'
    ws_location.cell(row=i, column=sched_id_idx+1, value=formula)

# Reorder sheets: Provider, ValidationAndReference, Location
sheet_order = ['Provider', 'ValidationAndReference', 'Location']
for idx, sheet_name in enumerate(sheet_order):
    if sheet_name in wb_out.sheetnames:
        wb_out.move_sheet(wb_out[sheet_name], offset=idx - wb_out.sheetnames.index(sheet_name))
wb_out.save(output_file)

# Add dropdown for 'Enterprise Scheduling Flag'
set_enterprise_scheduling_flag_dropdown(output_file)

print(f"File created with template structure at {output_file}.")

# Add specialty dropdowns to Provider sheet (after ValidationAndReference sheet is present)
# Remove or comment out the following lines:
# ws_valref = wb_out['ValidationAndReference']
# add_specialty_dropdowns(ws_out, ws_valref)

# Add Professional Suffix dropdowns at the end
from professional_suffix import add_professional_suffix_dropdowns
add_professional_suffix_dropdowns(output_file) 

# Add Hospital Affiliation dropdowns at the end
def add_hospital_affiliation_dropdowns(output_file):
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    wb = load_workbook(output_file)
    ws = wb['Provider']
    header_row = [cell.value for cell in ws[1]]
    for i in range(1, 6):
        col_name = f'Hospital Affiliation {i}'
        try:
            col_idx = header_row.index(col_name) + 1  # 1-based index
        except ValueError:
            continue  # Skip if the column is not found
        dv = DataValidation(type="list", formula1='=ValidationAndReference!$T$2:$T$7258', allow_blank=True)
        max_row = ws.max_row
        col_letter = get_column_letter(col_idx)
        dv_range = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(dv_range)
        ws.add_data_validation(dv)
    wb.save(output_file)

add_hospital_affiliation_dropdowns(output_file)

# Add Additional Langiage Spoken dropdowns at the end
from Langauge import set_additional_language_dropdowns
set_additional_language_dropdowns(output_file) 

# Add all specified dropdowns using provider_dropdowns.py
from provider_dropdowns import apply_provider_dropdowns, apply_provider_formulas

dropdown_specs = []
# Specialty 1-5
for i in range(1, 6):
    dropdown_specs.append((f"Specialty {i}", "=ValidationAndReference!$K$2:$K$311"))
# Board Certification 1-5
def add_board_certification_dropdowns():
    for i in range(1, 6):
        dropdown_specs.append((f"Board Certification {i}", "=ValidationAndReference!$N$2:$N$299"))
add_board_certification_dropdowns()
# Sub Board Certification 1-5
for i in range(1, 6):
    dropdown_specs.append((f"Sub Board Certification {i}", "=ValidationAndReference!$AB$2:$AB$156"))
# Additional Languages Spoken 1-5
for i in range(1, 6):
    dropdown_specs.append((f"Additional Languages Spoken {i}", "=ValidationAndReference!$W$2:$W$144"))
# Enterprise Scheduling Flag (only 'Yes')
dropdown_specs.append(("Enterprise Scheduling Flag", '"Yes"'))
# Practice Name dropdown from Location sheet
dropdown_specs.append(("Practice Name", "=Location!$A$8:$A$66"))
# Location 1-5 dropdowns from Location sheet
for i in range(1, 6):
    dropdown_specs.append((f"Location {i}", "=Location!$B$2:$B$236"))
# Provider Type dropdown from ValidationAndReference sheet
dropdown_specs.append(("Provider Type", "=ValidationAndReference!$Q$2:$Q$9"))

# Remove any previous 'Opt Out of Ratings' dropdown if present
# (No previous explicit entry, but ensure only 'Yes' is set)
dropdown_specs.append(("Opt Out of Ratings", '"Yes"'))

apply_provider_dropdowns(output_file, dropdown_specs)

# Apply formulas to specified columns
formula_specs = [
    ("Opt Out of Ratings", '=IFERROR(INDEX(ValidationAndReference!P:P,MATCH(BD{row},ValidationAndReference!Q:Q,0)),"")'),
    # Removed formula for 'Provider Type (Substatus) ID'
    ("Specialty ID 1", '=IF(ISBLANK(H{row}),"",INDEX(ValidationAndReference!J:J,MATCH(H{row},ValidationAndReference!K:K,0)))'),
    ("Specialty ID 2", '=IF(ISBLANK(H{row}),"",INDEX(ValidationAndReference!J:J,MATCH(I{row},ValidationAndReference!K:K,0)))'),
    ("Specialty ID 3", '=IF(ISBLANK(H{row}),"",INDEX(ValidationAndReference!J:J,MATCH(J{row},ValidationAndReference!K:K,0)))'),
    ("Specialty ID 4", '=IF(ISBLANK(H{row}),"",INDEX(ValidationAndReference!J:J,MATCH(K{row},ValidationAndReference!K:K,0)))'),
    ("Specialty ID 5", '=IF(ISBLANK(H{row}),"",INDEX(ValidationAndReference!J:J,MATCH(L{row},ValidationAndReference!K:K,0)))'),
    ("Location ID 1", '=IF(ISBLANK(N{row}),"",INDEX(Location!$U:$U,MATCH(N{row},Location!$A:$A,0)))'),
    ("Location ID 2", '=IF(ISBLANK(O{row}),"",INDEX(Location!$U:$U,MATCH(O{row},Location!$A:$A,0)))'),
    ("Location ID 3", '=IF(ISBLANK(P{row}),"",INDEX(Location!$U:$U,MATCH(P{row},Location!$A:$A,0)))'),
    ("Location ID 4", '=IF(ISBLANK(Q{row}),"",INDEX(Location!$U:$U,MATCH(Q{row},Location!$A:$A,0)))'),
    ("Location ID 5", '=IF(ISBLANK(R{row}),"",INDEX(Location!$U:$U,MATCH(R{row},Location!$A:$A,0)))'),
    ("Board Cert ID 1", '=IF(ISBLANK(AA{row}),"",INDEX(ValidationAndReference!$AA:$AA,MATCH(AA{row},ValidationAndReference!$AB:$AB,0)))'),
    ("Sub Board Cert ID 1", '=IF(ISBLANK(AB{row}),"",INDEX(ValidationAndReference!$M:$M,MATCH(AB{row},ValidationAndReference!$N:$N,0)))'),
    ("Board Cert ID 2", '=IF(ISBLANK(AC{row}),"",INDEX(ValidationAndReference!$AA:$AA,MATCH(AC{row},ValidationAndReference!$AB:$AB,0)))'),
    ("Sub Board Cert ID 2", '=IF(ISBLANK(AD{row}),"",INDEX(ValidationAndReference!$M:$M,MATCH(AD{row},ValidationAndReference!$N:$N,0)))'),
    ("Board Cert ID 3", '=IF(ISBLANK(AE{row}),"",INDEX(ValidationAndReference!$AA:$AA,MATCH(AE{row},ValidationAndReference!$AB:$AB,0)))'),
    ("Sub Board Cert ID 3", '=IF(ISBLANK(AF{row}),"",INDEX(ValidationAndReference!$M:$M,MATCH(AF{row},ValidationAndReference!$N:$N,0)))'),
    ("Board Cert ID 4", '=IF(ISBLANK(AG{row}),"",INDEX(ValidationAndReference!$AA:$AA,MATCH(AG{row},ValidationAndReference!$AB:$AB,0)))'),
    ("Sub Board Cert ID 4", '=IF(ISBLANK(AH{row}),"",INDEX(ValidationAndReference!$M:$M,MATCH(AH{row},ValidationAndReference!$N:$N,0)))'),
    ("Board Cert ID 5", '=IF(ISBLANK(AI{row}),"",INDEX(ValidationAndReference!$AA:$AA,MATCH(AI{row},ValidationAndReference!$AB:$AB,0)))'),
    ("Sub Board Cert ID 5", '=IF(ISBLANK(AJ{row}),"",INDEX(ValidationAndReference!$M:$M,MATCH(AJ{row},ValidationAndReference!$N:$N,0)))'),
    ("Hospital Affiliation ID 1", '=IF(ISBLANK(AU{row}),"",INDEX(ValidationAndReference!$S:$S,MATCH(AU{row},ValidationAndReference!$T:$T,0)))'),
    ("Hospital Affiliation ID 2", '=IF(ISBLANK(AV{row}),"",INDEX(ValidationAndReference!$S:$S,MATCH(AV{row},ValidationAndReference!$T:$T,0)))'),
    ("Hospital Affiliation ID 3", '=IF(ISBLANK(AW{row}),"",INDEX(ValidationAndReference!$S:$S,MATCH(AW{row},ValidationAndReference!$T:$T,0)))'),
    ("Hospital Affiliation ID 4", '=IF(ISBLANK(AX{row}),"",INDEX(ValidationAndReference!$S:$S,MATCH(AX{row},ValidationAndReference!$T:$T,0)))'),
    ("Hospital Affiliation ID 5", '=IF(ISBLANK(AY{row}),"",INDEX(ValidationAndReference!$S:$S,MATCH(AY{row},ValidationAndReference!$T:$T,0)))'),
    ("Language ID 1", '=IF(ISBLANK(AZ{row}),"",INDEX(ValidationAndReference!$V:$V,MATCH(AZ{row},ValidationAndReference!$W:$W,0)))'),
    ("Language ID 2", '=IF(ISBLANK(BA{row}),"",INDEX(ValidationAndReference!$V:$V,MATCH(BA{row},ValidationAndReference!$W:$W,0)))'),
    ("Language ID 3", '=IF(ISBLANK(BB{row}),"",INDEX(ValidationAndReference!$V:$V,MATCH(BB{row},ValidationAndReference!$W:$W,0)))'),
    ("Professional Suffix ID 1", '=IF(ISBLANK(D{row}),"",INDEX(ValidationAndReference!$F:$F,MATCH(D{row},ValidationAndReference!$G:$G,0)))'),
    ("Professional Suffix ID 2", '=IF(ISBLANK(E{row}),"",INDEX(ValidationAndReference!$F:$F,MATCH(E{row},ValidationAndReference!$G:$G,0)))'),
    ("Professional Suffix ID 3", '=IF(ISBLANK(F{row}),"",INDEX(ValidationAndReference!$F:$F,MATCH(F{row},ValidationAndReference!$G:$G,0)))'),
]

apply_provider_formulas(output_file, formula_specs)

create_npi_specialty_excel(
    r'C:\Users\dhruv.bhattacharjee\Desktop\PDO Data Transposition\Scope Conversion_Lifestance\Excel Files\Input.xlsx',
    r'C:\Users\dhruv.bhattacharjee\Desktop\PDO Data Transposition\Scope Conversion_Lifestance\Excel Files\Npi-specialty.xlsx'
)

#https://github.com/dhruv-bhattacharjee-zocdoc/Scope-Conversion_Lifestance

