import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# File paths
input_path = 'Excel Files/Input.xlsx'
merged_path = 'Excel Files/Mergedoutput.xlsx'

# Green highlight style
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

# Read provider telehealth/in-office info from Input.xlsx
input_df = pd.read_excel(input_path)
tele_col = None
for col in input_df.columns:
    if col.strip().lower() == 'telehealth or in-office or both':
        tele_col = col
        break
if not tele_col:
    raise Exception("Column 'Telehealth or In-Office or Both' not found in Input.xlsx")
telehealth_flags = input_df[tele_col].astype(str).str.strip().str.lower() == 'telehealth'

# Open Provider tab in Mergedoutput.xlsx
wb = openpyxl.load_workbook(merged_path)
ws = wb['Provider']
header = [cell.value for cell in ws[1]]
try:
    locid2_col = header.index('Location ID 2') + 1
except ValueError:
    raise Exception("'Location ID 2' column not found in Provider sheet.")

# Highlight where needed
highlighted_count = 0
for rownum, telehealth_flag in enumerate(telehealth_flags, start=2):
    if telehealth_flag:
        cell = ws.cell(row=rownum, column=locid2_col)
        if cell.value is not None and str(cell.value).strip() != '':
            cell.fill = green_fill
            highlighted_count += 1
wb.save(merged_path)
print(f"Telehealth check done: {highlighted_count} Location ID 2 cells highlighted green for Telehealth providers.")
