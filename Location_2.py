import pandas as pd
from fuzzywuzzy import fuzz
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import subprocess

# Load the Excel file
excel_path = 'Excel Files/Mergedoutput.xlsx'
xl = pd.ExcelFile(excel_path)
df = xl.parse('Provider')
loc = xl.parse('Location')

# Load the abbreviation mapping
abbr_path = 'Excel Files/C1 Street Suffix Abbreviations.xlsx'
mapping_df = pd.read_excel(abbr_path)
abbr_to_full = dict(zip(mapping_df['Commonly Used Street Suffix or Abbreviation'].astype(str).str.strip().str.lower(),
                        mapping_df['Postal Service Standard Suffix Abbreviation'].astype(str).str.strip().str.lower()))
full_to_abbr = dict(zip(mapping_df['Postal Service Standard Suffix Abbreviation'].astype(str).str.strip().str.lower(),
                        mapping_df['Commonly Used Street Suffix or Abbreviation'].astype(str).str.strip().str.lower()))

def replace_abbr_both_ways(address):
    words = str(address).split()
    replaced1 = [abbr_to_full.get(w.strip(',.').lower(), w) for w in words]
    replaced2 = [full_to_abbr.get(w.strip(',.').lower(), w) for w in words]
    return ' '.join(replaced1), ' '.join(replaced2)

def all_address_reprs(addr):
    # Return original, abbr->full, and full->abbr
    abbr_full, full_abbr = replace_abbr_both_ways(addr)
    return {str(addr).strip().lower(), abbr_full.strip().lower(), full_abbr.strip().lower()}

# Initialize result columns
loc_id_1 = []
loc_id_2 = []

for idx, prow in df.iterrows():
    p_addr_variants = all_address_reprs(prow['Facility Address']) if pd.notnull(prow['Facility Address']) else {''}
    p_city = str(prow['Facility City']).strip().lower() if pd.notnull(prow['Facility City']) else ''
    p_zip = str(prow['Facility Zip']).strip() if pd.notnull(prow['Facility Zip']) else ''
    p_state = str(prow['Facility State']).strip().lower() if pd.notnull(prow['Facility State']) else ''
    p_addr2 = str(prow['Address line 2']).strip().lower() if pd.notnull(prow['Address line 2']) else ''

    matches = []
    for _, lrow in loc.iterrows():
        l_addr_variants = all_address_reprs(lrow['Address line 1']) if pd.notnull(lrow['Address line 1']) else {''}
        l_city = str(lrow['City']).strip().lower() if pd.notnull(lrow['City']) else ''
        l_zip = str(lrow['ZIP Code']).strip() if pd.notnull(lrow['ZIP Code']) else ''
        l_state = str(lrow['State']).strip().lower() if pd.notnull(lrow['State']) else ''
        l_addr2 = str(lrow['Address line 2 (Office/Suite #)']).strip().lower() if 'Address line 2 (Office/Suite #)' in lrow and pd.notnull(lrow['Address line 2 (Office/Suite #)']) else ''
        score = 0
        addr_match = any(fuzz.partial_ratio(pa, la) >= 85 for pa in p_addr_variants for la in l_addr_variants)
        if addr_match:
            score += 1
        if p_city == l_city:
            score += 1
        if p_zip == l_zip:
            score += 1
        if p_state == l_state:
            score += 1
        if p_addr2:
            if p_addr2 == l_addr2:
                score += 1
        else:
            score += 1  # Bonus point if no address2 to match
        if score >= 4:
            matches.append((score, str(lrow['Location Cloud ID'])))
    # Extract top match for ID 1, remainder for ID 2
    if matches:
        matches_sorted = sorted(matches, key=lambda x: -x[0])
        loc_id_1.append(matches_sorted[0][1])
        loc_id_2.append(','.join([mid for _, mid in matches_sorted[1:]]))
    else:
        loc_id_1.append('')
        loc_id_2.append('')

df['Location ID 1'] = loc_id_1
df['Location ID 2'] = loc_id_2

# Save using pandas, then apply formatting using openpyxl
with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Provider', index=False)

# ---- Now handle the 'Both' logic and highlight if needed ----
input_path = 'Excel Files/Input.xlsx'
input_df = pd.read_excel(input_path)

# Map something (ProviderID, NPI, or row order) - for now, assume same order as df. Adjust if matching key is needed.
wbook = load_workbook(excel_path)
ws = wbook['Provider']

try:
    locid1_col = [cell.value for cell in ws[1]].index('Location ID 1') + 1
    locid2_col = [cell.value for cell in ws[1]].index('Location ID 2') + 1
except ValueError:
    locid1_col = None
    locid2_col = None

# Style for coloring red
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

for idx in range(len(df)):
    # Defensive in case of mismatched row numbers	n
    if idx >= len(input_df):
        continue
    tele_or_office = str(input_df.iloc[idx].get('Telehealth or In-Office or Both', '')).strip().lower()
    id1 = str(df.iloc[idx]['Location ID 1']).strip()
    id2 = str(df.iloc[idx]['Location ID 2']).strip()
    if tele_or_office == 'both':
        # Should have two unique, nonblank IDs
        if not id1 or not id2 or id1 == id2:
            ws.cell(row=idx+2, column=locid2_col).fill = red_fill

wbook.save(excel_path)

# Open workbook and get both sheets
wbook = load_workbook(excel_path)
ws_provider = wbook['Provider']
# To get the calculated cell values for 'Complete Location', reload workbook with data_only=True
wbook_dataonly = load_workbook(excel_path, data_only=True)
ws_location = wbook_dataonly['Location']

HEADER_ROW = 1
provider_headers = [cell.value for cell in ws_provider[HEADER_ROW]]
def get_or_create_col(ws, headers, name):
    try:
        return headers.index(name) + 1
    except ValueError:
        idx = ws.max_column + 1
        ws.cell(row=HEADER_ROW, column=idx, value=name)
        return idx

locid1_col = get_or_create_col(ws_provider, provider_headers, 'Location ID 1')
locid2_col = get_or_create_col(ws_provider, provider_headers, 'Location ID 2')
loc1_col = get_or_create_col(ws_provider, provider_headers, 'Location 1')
loc2_col = get_or_create_col(ws_provider, provider_headers, 'Location 2')

ld_headers = [cell.value for cell in ws_location[HEADER_ROW]]
try:
    loc_cloudid_col = ld_headers.index('Location Cloud ID') + 1  # W
    complete_loc_col = ld_headers.index('Complete Location') + 1 # X
except ValueError as e:
    raise ValueError('Location Cloud ID or Complete Location column not found in Location tab.')

# Build lookup with displayed values (data_only)
loc_cloud_to_full = {}
for row in ws_location.iter_rows(min_row=2, max_col=complete_loc_col, values_only=True):
    key = row[loc_cloudid_col-1]
    val = row[complete_loc_col-1]
    if key:
        loc_cloud_to_full[str(key).strip()] = val

maxrow = ws_provider.max_row
for i in range(2, maxrow+1):
    id1 = ws_provider.cell(row=i, column=locid1_col).value
    id2 = ws_provider.cell(row=i, column=locid2_col).value
    locval1 = loc_cloud_to_full.get(str(id1).strip(), '') if id1 else ''
    locval2 = ''
    if id2:
        id2list = [x.strip() for x in str(id2).split(',') if x.strip()]
        if len(id2list) == 1:
            locval2 = loc_cloud_to_full.get(id2list[0], '')
        elif len(id2list) > 1:
            locval2 = ' | '.join([str(loc_cloud_to_full.get(x, '') or '') for x in id2list])
    ws_provider.cell(row=i, column=loc1_col, value=locval1)
    ws_provider.cell(row=i, column=loc2_col, value=locval2)

wbook.save(excel_path)

# Now, populate Location 1 and Location 2 columns with Excel formulas
wbook = load_workbook(excel_path)
ws_provider = wbook['Provider']

loc1_col = get_or_create_col(ws_provider, provider_headers, 'Location 1')
loc2_col = get_or_create_col(ws_provider, provider_headers, 'Location 2')

maxrow = ws_provider.max_row
for row in range(2, maxrow+1):
    # Row numbers in formulas must match the current Excel row
    formula_loc1 = '=IFERROR(INDEX(Location!X:X, MATCH(BR{row}, Location!W:W, 0)), "")'.format(row=row)
    formula_loc2 = '=IFERROR(INDEX(Location!X:X, MATCH(BS{row}, Location!W:W, 0)), "")'.format(row=row)
    ws_provider.cell(row=row, column=loc1_col, value=formula_loc1)
    ws_provider.cell(row=row, column=loc2_col, value=formula_loc2)

wbook.save(excel_path)

# Add requested columns and formulas to Provider tab
formula_targets = [
    ('Provider Type (Substatus) ID', '=IFERROR(INDEX(ValidationAndReference!P:P, MATCH(BE{row}, ValidationAndReference!Q:Q, 0)), "")'),
    ('Specialty 1', '=IFERROR(VLOOKUP(BM{row}, ValidationAndReference!J:K, 2, FALSE), "")'),
    ('Professional Suffix ID 1', '=IFERROR(INDEX(ValidationAndReference!F:F, MATCH(D{row}, ValidationAndReference!G:G, 0)), "")'),
    ('Professional Suffix ID 2', '=IFERROR(INDEX(ValidationAndReference!F:F, MATCH(E{row}, ValidationAndReference!G:G, 0)), "")'),
    ('Professional Suffix ID 3', '=IFERROR(INDEX(ValidationAndReference!F:F, MATCH(F{row}, ValidationAndReference!G:G, 0)), "")'),
    ('Language ID 1', '=IFERROR(INDEX(ValidationAndReference!V:V, MATCH(AZ{row}, ValidationAndReference!W:W, 0)), "")'),
    ('Language ID 2', '=IFERROR(INDEX(ValidationAndReference!V:V, MATCH(BA{row}, ValidationAndReference!W:W, 0)), "")'),
    ('Language ID 3', '=IFERROR(INDEX(ValidationAndReference!V:V, MATCH(BB{row}, ValidationAndReference!W:W, 0)), "")'),
]
# Reload headers in case more columns were inserted
provider_headers = [cell.value for cell in ws_provider[HEADER_ROW]]
col_indices = {}
for colname, _ in formula_targets:
    col_indices[colname] = get_or_create_col(ws_provider, provider_headers, colname)
# Populate each formula in each row
for row in range(2, ws_provider.max_row+1):
    for colname, formula_template in formula_targets:
        col = col_indices[colname]
        ws_provider.cell(row=row, column=col, value=formula_template.format(row=row))
wbook.save(excel_path)

# Remove values from certain columns when 'NPI Number' is blank
cols_to_blank = [
    'Patients Accepted',
    'Provider Type',
    'Enterprise Scheduling Flag',
    'Provider Type (Substatus) ID',
    'Matched',
]

npi_col_idx = None
# Refresh headers in case new columns added
provider_headers = [cell.value for cell in ws_provider[HEADER_ROW]]
try:
    npi_col_idx = provider_headers.index('NPI Number') + 1
except ValueError:
    npi_col_idx = None
col_indices = {}
for cname in cols_to_blank:
    try:
        col_indices[cname] = provider_headers.index(cname) + 1
    except ValueError:
        continue  # Don't error if a column doesn't exist
if npi_col_idx:
    for i in range(2, ws_provider.max_row+1):
        npi_val = ws_provider.cell(row=i, column=npi_col_idx).value
        if not npi_val or str(npi_val).strip() == '':
            for cname, cidx in col_indices.items():
                ws_provider.cell(row=i, column=cidx, value=None)
wbook.save(excel_path)

# Call the practice_check.py script as the last step
try:
    subprocess.run(['python', 'practice_check.py'], check=True)
    print("practice_check.py completed successfully.")
except Exception as e:
    print(f"Error running practice_check.py: {e}")

if __name__ == "__main__":
    # Call Telehealthcheck.py as the last step
    import subprocess
    print("Running Telehealthcheck.py as final step...")
    subprocess.run(['python', 'Telehealthcheck.py'], check=True)
    # Ensure file is saved after Telehealthcheck
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    wb.save(excel_path)
