import openpyxl
from typing import List, Optional
import re

def map_ages_to_patients_accepted(ages_treated: Optional[str]) -> str:
    if not ages_treated or not isinstance(ages_treated, str) or ages_treated.strip() == '':
        return 'Adult'
    # Split by comma and strip spaces
    age_ranges = [part.strip() for part in ages_treated.split(',') if part.strip()]
    has_pediatric = False
    has_adult = False
    for age_range in age_ranges:
        # Match single age or range
        match = re.match(r'^(\d+)(?:-(\d+))?$', age_range)
        if not match:
            continue
        start = int(match.group(1))
        end = int(match.group(2)) if match.group(2) else start
        # Pediatric: any part <= 17
        if start <= 17 or end <= 17:
            has_pediatric = True
        # Adult: any part >= 19
        if start >= 19 or end >= 19:
            has_adult = True
        # 18: both
        if start <= 18 <= end:
            has_pediatric = True
            has_adult = True
    if has_pediatric and has_adult:
        return 'Both'
    elif has_pediatric:
        return 'Pediatric'
    elif has_adult:
        return 'Adult'
    else:
        return 'Adult'

def extract_patients_accepted(input_file: str) -> List[Optional[str]]:
    """
    Extracts the 'Ages Treated' column from the input Excel file and returns a list of 'Adult', 'Pediatric', or 'Both'.
    """
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    if ws is None:
        raise Exception("No active worksheet found in the input file.")
    header = [str(cell.value) if cell.value is not None else '' for cell in ws[1]]
    try:
        ages_treated_idx = header.index('Ages Treated')
    except ValueError:
        raise Exception("'Ages Treated' column not found in input file.")
    patients_accepted = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is not None and len(row) > ages_treated_idx:
            ages_val = row[ages_treated_idx]
            patients_accepted.append(map_ages_to_patients_accepted(str(ages_val)) if ages_val is not None else 'Adult')
        else:
            patients_accepted.append('Adult')
    return patients_accepted

def set_patients_accepted_dropdown(output_file: str):
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Provider']
    header_row = [cell.value for cell in ws[1]]
    try:
        col_idx = header_row.index('Patients Accepted') + 1  # 1-based index
    except ValueError:
        raise Exception('Patients Accepted column not found in output file.')
    dv = DataValidation(type="list", formula1='"Adult,Pediatric,Both"', allow_blank=True)
    max_row = ws.max_row
    col_letter = get_column_letter(col_idx)
    dv_range = f"{col_letter}2:{col_letter}{max_row}"
    dv.add(dv_range)
    ws.add_data_validation(dv)
    wb.save(output_file)
