import pandas as pd
import openpyxl
import os
import re
from fuzzywuzzy import fuzz

# File paths
mapping_file = os.path.join('Excel Files', 'C1 Street Suffix Abbreviations.xlsx')
merged_file = os.path.join('Excel Files', 'Mergedoutput.xlsx')

# Load the mapping from the abbreviations file
mapping_df = pd.read_excel(mapping_file)

# Create mapping dictionary
# Strip and lower for robust matching
mapping_dict = dict(zip(
    mapping_df['Commonly Used Street Suffix or Abbreviation'].astype(str).str.strip().str.lower(),
    mapping_df['Postal Service Standard Suffix Abbreviation'].astype(str).str.strip()
))

# Camel case function
def to_camel_case(s):
    parts = s.split()
    return ''.join(word.capitalize() for word in parts)

# Helper for smart camel case (preserve NW, SW, SE, NE)
def smart_camel_case(text, keep_upper=None):
    if not isinstance(text, str):
        return text
    if keep_upper is None:
        keep_upper = {'NE', 'SE', 'SW', 'NW', 'N', 'S', 'E', 'W'}
    def fix_word(word):
        w = word.strip()
        if w.upper() in keep_upper:
            return w.upper()
        return w[:1].upper() + w[1:].lower() if w else w
    def fix_token(token):
        return '-'.join(fix_word(part) for part in token.split('-'))
    return ' '.join(fix_token(token) for token in text.split())

# Suite/unit/PO Box regex (same as Location.py)
suite_keywords = [
    r"suite", r"ste", r"apt", r"apartment", r"floor", r"fl", r"unit", r"room", r"rm", r"bldg", r"#", r"p\.o\. box", r"po box"
]
suite_pattern = re.compile(r"(" + r"|".join(suite_keywords) + r").*", re.IGNORECASE)

# Hardcoded exceptions for common abbreviations
exception_abbr = {
    'floor': 'FL',
    'suite': 'STE',
}

def split_and_clean_address(address):
    """
    Splits an address into main and secondary parts (suite/unit/PO Box or after comma),
    applies suffix mapping to all words of both address_line_1 and address_line_2, then camel cases both (preserving NW, SW, SE, NE), and returns them as (address_line_1, address_line_2).
    """
    address_line_1 = address if address else ""
    address_line_2 = ""
    if address_line_1:
        addr_str = str(address_line_1)
        # If there's a comma, move everything after the first comma to Address line 2
        if ',' in addr_str:
            before_comma, after_comma = addr_str.split(',', 1)
            address_line_1 = before_comma.strip()
            address_line_2 = after_comma.strip()
        # Move suite/unit/PO Box info to Address line 2
        match = suite_pattern.search(address_line_1)
        if match:
            suite_part = address_line_1[match.start():].strip()
            address_line_1 = address_line_1[:match.start()].strip()
            if address_line_2:
                address_line_2 = f"{address_line_2} {suite_part}".strip()
            else:
                address_line_2 = suite_part
        # --- Apply suffix mapping to all words of address_line_1 and address_line_2 BEFORE camel casing ---
        def apply_suffix_mapping(addr):
            words = addr.split()
            new_words = []
            for word in words:
                base_word = word.strip(',').strip('.')
                key = base_word.lower()
                abbr = None
                if key in exception_abbr:
                    abbr = exception_abbr[key]
                elif key in mapping_dict:
                    abbr = mapping_dict[key]
                if abbr:
                    # Keep any trailing comma or period
                    suffix = ''
                    if word.endswith(','):
                        suffix = ','
                    elif word.endswith('.'):
                        suffix = '.'
                    new_words.append(abbr + suffix)
                else:
                    new_words.append(word)
            return ' '.join(new_words)
        # Apply mapping first, then camel case
        address_line_1 = smart_camel_case(apply_suffix_mapping(address_line_1))
        address_line_2 = smart_camel_case(apply_suffix_mapping(address_line_2)) if address_line_2 else ""
    return address_line_1, address_line_2

# Load the Provider sheet from the merged file
with pd.ExcelFile(merged_file) as xls:
    # Read all sheets
    sheets = {sheet: xls.parse(sheet) for sheet in xls.sheet_names}

# Check if 'Provider' sheet and 'Facility Address' column exist
if 'Provider' not in sheets:
    raise ValueError("'Provider' sheet not found in Mergedoutput.xlsx")

provider_df = sheets['Provider']
if 'Facility Address' not in provider_df.columns:
    raise ValueError("'Facility Address' column not found in Provider sheet")

# Function to replace suffixes in address
def replace_suffix(address):
    if not isinstance(address, str):
        return address
    words = address.split()
    new_words = []
    for word in words:
        key = word.strip(',').lower()
        # Replace if mapping exists (preserve punctuation)
        if key in mapping_dict:
            # Keep any trailing comma or period
            suffix = ''
            if word.endswith(','):
                suffix = ','
            elif word.endswith('.'):
                suffix = '.'
            camel_cased = to_camel_case(mapping_dict[key])
            new_word = camel_cased + suffix
            new_words.append(new_word)
        else:
            new_words.append(word)
    return ' '.join(new_words)

# Apply the replacement
provider_df['Facility Address'] = provider_df['Facility Address'].apply(replace_suffix)

# --- Provider/Location matching logic ---
if 'Location' not in sheets:
    raise ValueError("'Location' sheet not found in Mergedoutput.xlsx")
location_df = sheets['Location']

# --- Remove duplicates in Location sheet based on 'Location Cloud ID' (keep all rows where this column is nan/blank) ---
if 'Location Cloud ID' in location_df.columns:
    # Identify rows where 'Location Cloud ID' is not null/blank
    mask_notnull = location_df['Location Cloud ID'].notna() & (location_df['Location Cloud ID'].astype(str).str.strip() != '')
    # Split into rows with and without valid Location Cloud ID
    df_with_id = location_df[mask_notnull]
    df_without_id = location_df[~mask_notnull]
    # Drop duplicates in rows with valid Location Cloud ID, keeping the first occurrence
    df_with_id_nodup = df_with_id.drop_duplicates(subset=['Location Cloud ID'], keep='first')
    # Concatenate back, keeping all rows with nan/blank
    location_df = pd.concat([df_with_id_nodup, df_without_id], ignore_index=True)
    # Optional: sort to keep original order as much as possible (not strictly necessary)
    location_df = location_df.sort_index(kind='stable')
    # Overwrite the Location sheet in the file
    with pd.ExcelWriter(merged_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        location_df.to_excel(writer, sheet_name='Location', index=False)

    # --- Apply formula to 'Complete Location' column using openpyxl ---
    from openpyxl import load_workbook
    wb = load_workbook(merged_file)
    ws = wb['Location']
    # Find or create the 'Complete Location' column
    header = [cell.value for cell in ws[1]]
    try:
        complete_loc_col_idx = header.index('Complete Location') + 1  # 1-based
    except ValueError:
        complete_loc_col_idx = ws.max_column + 1
        ws.cell(row=1, column=complete_loc_col_idx, value='Complete Location')
    # Apply the formula to each row (starting from row 2)
    for row in range(2, ws.max_row + 1):
        formula = f'=IF(A{row}<>"",CONCATENATE(A{row}," ",B{row}," ",D{row}," ",E{row}," ",F{row}," ",G{row}," ","(",C{row},")"),"")'
        ws.cell(row=row, column=complete_loc_col_idx, value=formula)
    wb.save(merged_file)

    # --- Re-apply yellow fill to entire row where 'Location Cloud ID' is blank/NaN ---
    from openpyxl.styles import PatternFill
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    # Find the 'Location Cloud ID' column index
    header = [cell.value for cell in ws[1]]
    try:
        loc_cloud_id_col_idx = header.index('Location Cloud ID') + 1  # 1-based
    except ValueError:
        loc_cloud_id_col_idx = None
    if loc_cloud_id_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=loc_cloud_id_col_idx)
            if cell.value is None or str(cell.value).strip() == '':
                # Highlight the entire row
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill
    wb.save(merged_file)

# Ensure columns exist
required_provider_cols = ['Facility Address', 'Facility City', 'Facility Zip', 'Facility State', 'Address line 2']
required_location_cols = ['Address line 1', 'City', 'ZIP Code', 'State', 'Address line 2 (Office/Suite #)']
for col in required_provider_cols:
    if col not in provider_df.columns:
        raise ValueError(f"'{col}' column not found in Provider sheet")
for col in required_location_cols:
    if col not in location_df.columns:
        raise ValueError(f"'{col}' column not found in Location sheet")

# --- Now create the 'Matched' column in Provider sheet ---
matched_results = []
location_id_1_results = []  # For 'In Person'
location_id_2_results = []  # For 'Virtual'
for idx, prow in provider_df.iterrows():
    reasons = []
    suggestions = []
    match_type = None
    abbr_reason = None
    # Prepare values (case-insensitive, strip)
    p_addr = str(prow['Facility Address']).strip().lower() if pd.notnull(prow['Facility Address']) else ''
    p_city = str(prow['Facility City']).strip().lower() if pd.notnull(prow['Facility City']) else ''
    p_zip = str(prow['Facility Zip']).strip() if pd.notnull(prow['Facility Zip']) else ''
    p_state = str(prow['Facility State']).strip().lower() if pd.notnull(prow['Facility State']) else ''
    p_addr2 = str(prow['Address line 2']).strip().lower() if pd.notnull(prow['Address line 2']) else ''
    # Fuzzy match for address
    def fuzzy_addr_match(addr):
        return any(fuzz.partial_ratio(addr, str(loc_addr).strip().lower()) >= 85 for loc_addr in location_df['Address line 1'])
    addr_match_mask = location_df['Address line 1'].apply(lambda x: fuzz.partial_ratio(p_addr, str(x).strip().lower()) >= 85)
    mask = (
        addr_match_mask &
        (location_df['City'].astype(str).str.strip().str.lower() == p_city) &
        (location_df['ZIP Code'].astype(str).str.strip() == p_zip) &
        (location_df['State'].astype(str).str.strip().str.lower() == p_state)
    )
    if p_addr2:
        mask = mask & (location_df['Address line 2 (Office/Suite #)'].astype(str).str.strip().str.lower() == p_addr2)
    match = location_df[mask]
    # Direct fuzzy match (fuzzywuzzy, not abbreviation)
    direct_fuzzy_match = False
    if not match.empty:
        # Check if it's a direct substring match (not just fuzzy)
        direct_substring = location_df['Address line 1'].astype(str).str.lower().str.contains(p_addr, na=False)
        direct_substring_mask = (
            direct_substring &
            (location_df['City'].astype(str).str.strip().str.lower() == p_city) &
            (location_df['ZIP Code'].astype(str).str.strip() == p_zip) &
            (location_df['State'].astype(str).str.strip().str.lower() == p_state)
        )
        if p_addr2:
            direct_substring_mask = direct_substring_mask & (location_df['Address line 2 (Office/Suite #)'].astype(str).str.strip().str.lower() == p_addr2)
        if location_df[direct_substring_mask].empty:
            direct_fuzzy_match = True
    # If not matched, check for possible suffix/abbreviation mapping
    is_suffix_match = False
    abbr_word = None
    abbr_replacement = None
    if match.empty:
        words = [w.strip(',').strip('.') for w in prow['Facility Address'].split()] if pd.notnull(prow['Facility Address']) else []
        for word in words:
            # Try both directions: Commonly Used <-> Postal Service
            # 1. Commonly Used -> Postal Service
            matches = mapping_df[mapping_df['Commonly Used Street Suffix or Abbreviation'].astype(str).str.lower() == word.lower()]
            for _, row in matches.iterrows():
                alt_addr = p_addr.replace(word, str(row['Postal Service Standard Suffix Abbreviation']).lower())
                alt_addr_mask = location_df['Address line 1'].apply(lambda x: fuzz.partial_ratio(alt_addr, str(x).strip().lower()) >= 85)
                alt_mask = (
                    alt_addr_mask &
                    (location_df['City'].astype(str).str.strip().str.lower() == p_city) &
                    (location_df['ZIP Code'].astype(str).str.strip() == p_zip) &
                    (location_df['State'].astype(str).str.strip().str.lower() == p_state)
                )
                if p_addr2:
                    alt_mask = alt_mask & (location_df['Address line 2 (Office/Suite #)'].astype(str).str.strip().str.lower() == p_addr2)
                alt_match = location_df[alt_mask]
                if not alt_match.empty:
                    is_suffix_match = True
                    abbr_word = word
                    abbr_replacement = str(row['Postal Service Standard Suffix Abbreviation'])
                    break
            if is_suffix_match:
                break
            # 2. Postal Service -> Commonly Used
            matches2 = mapping_df[mapping_df['Postal Service Standard Suffix Abbreviation'].astype(str).str.lower() == word.lower()]
            for _, row in matches2.iterrows():
                alt_addr = p_addr.replace(word, str(row['Commonly Used Street Suffix or Abbreviation']).lower())
                alt_addr_mask = location_df['Address line 1'].apply(lambda x: fuzz.partial_ratio(alt_addr, str(x).strip().lower()) >= 85)
                alt_mask = (
                    alt_addr_mask &
                    (location_df['City'].astype(str).str.strip().str.lower() == p_city) &
                    (location_df['ZIP Code'].astype(str).str.strip() == p_zip) &
                    (location_df['State'].astype(str).str.strip().str.lower() == p_state)
                )
                if p_addr2:
                    alt_mask = alt_mask & (location_df['Address line 2 (Office/Suite #)'].astype(str).str.strip().str.lower() == p_addr2)
                alt_match = location_df[alt_mask]
                if not alt_match.empty:
                    is_suffix_match = True
                    abbr_word = word
                    abbr_replacement = str(row['Commonly Used Street Suffix or Abbreviation'])
                    break
            if is_suffix_match:
                break
    # Set Matched column with reason if not direct substring match
    if not match.empty:
        if direct_fuzzy_match:
            matched_results.append('Yes {fuzzy match}')
        else:
            matched_results.append('Yes')
        # --- Location ID logic ---
        in_person_match = match[match['Location Type'].astype(str).str.strip().str.lower() == 'in person']
        virtual_match = match[match['Location Type'].astype(str).str.strip().str.lower() == 'virtual']
        location_id_1 = in_person_match.iloc[0]['Location Cloud ID'] if not in_person_match.empty else ''
        location_id_2 = virtual_match.iloc[0]['Location Cloud ID'] if not virtual_match.empty else ''
        location_id_1_results.append(location_id_1)
        location_id_2_results.append(location_id_2)
        continue
    elif is_suffix_match:
        matched_results.append(f"Yes {{abbreviation match: '{abbr_word}' to '{abbr_replacement}'}}")
        # Try to find Location IDs for abbreviation match as well
        alt_addr = p_addr.replace(abbr_word.lower(), abbr_replacement.lower())
        alt_addr_mask = location_df['Address line 1'].apply(lambda x: fuzz.partial_ratio(alt_addr, str(x).strip().lower()) >= 85)
        alt_mask = (
            alt_addr_mask &
            (location_df['City'].astype(str).str.strip().str.lower() == p_city) &
            (location_df['ZIP Code'].astype(str).str.strip() == p_zip) &
            (location_df['State'].astype(str).str.strip().str.lower() == p_state)
        )
        if p_addr2:
            alt_mask = alt_mask & (location_df['Address line 2 (Office/Suite #)'].astype(str).str.strip().str.lower() == p_addr2)
        alt_match = location_df[alt_mask]
        in_person_match = alt_match[alt_match['Location Type'].astype(str).str.strip().str.lower() == 'in person']
        virtual_match = alt_match[alt_match['Location Type'].astype(str).str.strip().str.lower() == 'virtual']
        location_id_1 = in_person_match.iloc[0]['Location Cloud ID'] if not in_person_match.empty else ''
        location_id_2 = virtual_match.iloc[0]['Location Cloud ID'] if not virtual_match.empty else ''
        location_id_1_results.append(location_id_1)
        location_id_2_results.append(location_id_2)
        continue
    # If still not matched, report reasons and suggestions
    addr_match = fuzzy_addr_match(p_addr) or is_suffix_match
    city_match = (location_df['City'].astype(str).str.strip().str.lower() == p_city).any()
    zip_match = (location_df['ZIP Code'].astype(str).str.strip() == p_zip).any()
    state_match = (location_df['State'].astype(str).str.strip().str.lower() == p_state).any()
    if not addr_match:
        # Try to suggest suffix/abbreviation fixes
        words = [w.strip(',').strip('.') for w in prow['Facility Address'].split()] if pd.notnull(prow['Facility Address']) else []
        for word in words:
            matches = mapping_df[mapping_df['Commonly Used Street Suffix or Abbreviation'].astype(str).str.lower() == word.lower()]
            for _, row in matches.iterrows():
                suggestions.append(f"try converting '{word}' to '{row['Postal Service Standard Suffix Abbreviation']}'")
            matches2 = mapping_df[mapping_df['Postal Service Standard Suffix Abbreviation'].astype(str).str.lower() == word.lower()]
            for _, row in matches2.iterrows():
                suggestions.append(f"try converting '{word}' to '{row['Commonly Used Street Suffix or Abbreviation']}'")
        if suggestions:
            reasons.append('Facility Address (' + '; '.join(suggestions) + ')')
        else:
            reasons.append('Facility Address')
    if not city_match:
        reasons.append('Facility City')
    if not zip_match:
        reasons.append('Facility Zip')
    if not state_match:
        reasons.append('Facility State')
    if p_addr2:
        addr2_match = (location_df['Address line 2 (Office/Suite #)'].astype(str).str.strip().str.lower() == p_addr2).any()
        if not addr2_match:
            reasons.append('Address line 2')
    matched_results.append('No - ' + ', '.join(reasons))
    location_id_1_results.append('')
    location_id_2_results.append('')
provider_df['Matched'] = matched_results
provider_df['Location ID 1'] = location_id_1_results
provider_df['Location ID 2'] = location_id_2_results

# Fill 'Location 1' and 'Location 2' columns based on 'Location ID 1' and 'Location ID 2'
location_1_results = []
location_2_results = []
for idx, row in provider_df.iterrows():
    loc_id_1 = row['Location ID 1']
    loc_id_2 = row['Location ID 2']
    loc_1 = ''
    loc_2 = ''
    if loc_id_1:
        match_row = location_df[location_df['Location Cloud ID'] == loc_id_1]
        if not match_row.empty and 'Complete Location' in match_row.columns:
            loc_1 = match_row.iloc[0]['Complete Location']
    if loc_id_2:
        match_row = location_df[location_df['Location Cloud ID'] == loc_id_2]
        if not match_row.empty and 'Complete Location' in match_row.columns:
            loc_2 = match_row.iloc[0]['Complete Location']
    location_1_results.append(loc_1)
    location_2_results.append(loc_2)
provider_df['Location 1'] = location_1_results
provider_df['Location 2'] = location_2_results

# Post-process: If 'Location ID 1' is empty but 'Location ID 2' is not, move 'Location ID 2' to 'Location ID 1' and clear 'Location ID 2'
for idx, row in provider_df.iterrows():
    if (not row['Location ID 1']) and row['Location ID 2']:
        provider_df.at[idx, 'Location ID 1'] = row['Location ID 2']
        provider_df.at[idx, 'Location ID 2'] = ''

# Overwrite the Provider sheet in the original file
with pd.ExcelWriter(merged_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    provider_df.to_excel(writer, sheet_name='Provider', index=False)

# Highlight empty 'Location ID 1' cells in yellow
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
wb = load_workbook(merged_file)
ws = wb['Provider']
header = [cell.value for cell in ws[1]]
try:
    loc_id_1_col_idx = header.index('Location ID 1') + 1  # 1-based
except ValueError:
    loc_id_1_col_idx = None
if loc_id_1_col_idx:
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=loc_id_1_col_idx)
        if cell.value is None or str(cell.value).strip() == '':
            cell.fill = yellow_fill
wb.save(merged_file)

# Utility functions (split_and_clean_address, etc.) remain at top level

def main():
    # Load the Provider sheet from the output file (now using Output.xlsx)
    output_file = os.path.join('Excel Files', 'Output.xlsx')
    mapping_file = os.path.join('Excel Files', 'C1 Street Suffix Abbreviations.xlsx')
    mapping_df = pd.read_excel(mapping_file)
    mapping_dict = dict(zip(
        mapping_df['Commonly Used Street Suffix or Abbreviation'].astype(str).str.strip().str.lower(),
        mapping_df['Postal Service Standard Suffix Abbreviation'].astype(str).str.strip()
    ))
    with pd.ExcelFile(output_file) as xls:
        sheets = {sheet: xls.parse(sheet) for sheet in xls.sheet_names}
    if 'Provider' not in sheets:
        raise ValueError("'Provider' sheet not found in Output.xlsx")
    provider_df = sheets['Provider']
    if 'Facility Address' not in provider_df.columns:
        raise ValueError("'Facility Address' column not found in Provider sheet")
    if 'Location' not in sheets:
        raise ValueError("'Location' sheet not found in Output.xlsx")
    location_df = sheets['Location']

    # --- Remove duplicates in Location sheet based on 'Location Cloud ID' (keep all rows where this column is nan/blank) ---
    if 'Location Cloud ID' in location_df.columns:
        # Identify rows where 'Location Cloud ID' is not null/blank
        mask_notnull = location_df['Location Cloud ID'].notna() & (location_df['Location Cloud ID'].astype(str).str.strip() != '')
        # Split into rows with and without valid Location Cloud ID
        df_with_id = location_df[mask_notnull]
        df_without_id = location_df[~mask_notnull]
        # Drop duplicates in rows with valid Location Cloud ID, keeping the first occurrence
        df_with_id_nodup = df_with_id.drop_duplicates(subset=['Location Cloud ID'], keep='first')
        # Concatenate back, keeping all rows with nan/blank
        location_df = pd.concat([df_with_id_nodup, df_without_id], ignore_index=True)
        # Optional: sort to keep original order as much as possible (not strictly necessary)
        location_df = location_df.sort_index(kind='stable')
        # Overwrite the Location sheet in the file
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            location_df.to_excel(writer, sheet_name='Location', index=False)

    # --- Apply formula to 'Complete Location' column using openpyxl ---
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    ws = wb['Location']
    # Find or create the 'Complete Location' column
    header = [cell.value for cell in ws[1]]
    try:
        complete_loc_col_idx = header.index('Complete Location') + 1  # 1-based
    except ValueError:
        complete_loc_col_idx = ws.max_column + 1
        ws.cell(row=1, column=complete_loc_col_idx, value='Complete Location')
    # Apply the formula to each row (starting from row 2)
    for row in range(2, ws.max_row + 1):
        formula = f'=IF(A{row}<>"",CONCATENATE(A{row}," ",B{row}," ",D{row}," ",E{row}," ",F{row}," ",G{row}," ","(",C{row},")"),"")'
        ws.cell(row=row, column=complete_loc_col_idx, value=formula)
    wb.save(output_file)

    # --- Re-apply yellow fill to entire row where 'Location Cloud ID' is blank/NaN ---
    from openpyxl.styles import PatternFill
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    # Find the 'Location Cloud ID' column index
    header = [cell.value for cell in ws[1]]
    try:
        loc_cloud_id_col_idx = header.index('Location Cloud ID') + 1  # 1-based
    except ValueError:
        loc_cloud_id_col_idx = None
    if loc_cloud_id_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=loc_cloud_id_col_idx)
            if cell.value is None or str(cell.value).strip() == '':
                # Highlight the entire row
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill
    wb.save(output_file)

    required_provider_cols = ['Facility Address', 'Facility City', 'Facility Zip', 'Facility State', 'Address line 2']
    required_location_cols = ['Address line 1', 'City', 'ZIP Code', 'State', 'Address line 2 (Office/Suite #)']
    for col in required_provider_cols:
        if col not in provider_df.columns:
            raise ValueError(f"'{col}' column not found in Provider sheet")
    for col in required_location_cols:
        if col not in location_df.columns:
            raise ValueError(f"'{col}' column not found in Location sheet")
    # ... (rest of the matching logic goes here, unchanged, but using output_file instead of merged_file) ...
    # (Copy the entire matching logic block here, replacing merged_file with output_file)
    # Overwrite the Provider sheet in the original file
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        provider_df.to_excel(writer, sheet_name='Provider', index=False)

if __name__ == '__main__':
    main()
