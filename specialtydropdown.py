import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

def add_specialty_valref_dropdowns(output_file):
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Provider']
    header_row = [cell.value for cell in ws[1]]
    for col_name in [f'Specialty {i}' for i in range(2, 6)]:
        try:
            col_idx = header_row.index(col_name) + 1  # 1-based index for openpyxl
        except ValueError:
            continue  # Skip if the column is not found
        dv = DataValidation(type="list", formula1='=ValidationAndReference!$K$2:$K$311', allow_blank=True)
        max_row = ws.max_row
        col_letter = get_column_letter(col_idx)
        dv_range = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(dv_range)
        ws.add_data_validation(dv)
    wb.save(output_file)

if __name__ == "__main__":
    output_file = os.path.join("Excel Files", "Output.xlsx")
    add_specialty_valref_dropdowns(output_file) 