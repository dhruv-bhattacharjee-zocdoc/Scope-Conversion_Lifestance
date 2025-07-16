import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def extract_languages(input_file):
    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in.active
    if ws_in is None:
        raise ValueError("Input worksheet could not be loaded.")
    input_header_row = [cell.value for cell in ws_in[1]]
    try:
        lang_idx = input_header_row.index('Languages')
    except ValueError:
        raise ValueError("'Languages' column not found in input file.")
    lang1_list = []
    lang2_list = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        cell = row[lang_idx]
        if cell is not None and str(cell).strip() != '':
            parts = [part.strip() for part in str(cell).split(',')]
            lang1_list.append(parts[0] if len(parts) > 0 else "")
            lang2_list.append(parts[1] if len(parts) > 1 else "")
        else:
            lang1_list.append("")
            lang2_list.append("")
    return lang1_list, lang2_list

def set_additional_language_dropdowns(output_file):
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Provider']
    header_row = [cell.value for cell in ws[1]]
    for i in range(1, 4):
        col_name = f'Additional Langiage Spoken {i}'
        try:
            col_idx = header_row.index(col_name) + 1  # 1-based index
        except ValueError:
            continue  # Skip if the column is not found
        dv = DataValidation(type="list", formula1='=ValidationAndReference!$W$2:$W$144', allow_blank=True)
        max_row = ws.max_row
        col_letter = get_column_letter(col_idx)
        dv_range = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(dv_range)
        ws.add_data_validation(dv)
    wb.save(output_file) 