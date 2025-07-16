import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def set_enterprise_scheduling_flag_dropdown(output_file):
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Provider']
    # Find the column index for 'Enterprise Scheduling Flag'
    header_row = [cell.value for cell in ws[1]]
    try:
        col_idx = header_row.index('Enterprise Scheduling Flag') + 1  # 1-based index for openpyxl
    except ValueError:
        raise ValueError("'Enterprise Scheduling Flag' column not found in output file.")
    # Create the data validation dropdown
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    # The range should cover all rows in the column except the header
    max_row = ws.max_row
    col_letter = get_column_letter(col_idx)
    dv_range = f"{col_letter}2:{col_letter}{max_row}"
    dv.add(dv_range)
    ws.add_data_validation(dv)
    wb.save(output_file)
