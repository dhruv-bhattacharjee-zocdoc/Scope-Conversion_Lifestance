import openpyxl

def extract_board_certification(input_file):
    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in.active
    if ws_in is None:
        raise ValueError("Input worksheet could not be loaded.")
    input_header_row = [cell.value for cell in ws_in[1]]
    try:
        cert_idx = input_header_row.index('Board Certification')
    except ValueError:
        raise ValueError("'Board Certification' column not found in input file.")
    cert_list = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        cert_list.append(row[cert_idx])
    return cert_list

def extract_board_subspecialty(input_file):
    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in.active
    if ws_in is None:
        raise ValueError("Input worksheet could not be loaded.")
    input_header_row = [cell.value for cell in ws_in[1]]
    try:
        subspecialty_idx = input_header_row.index('Board Subspecialty')
    except ValueError:
        raise ValueError("'Board Subspecialty' column not found in input file.")
    subspecialty_list = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        subspecialty_list.append(row[subspecialty_idx])
    return subspecialty_list

def set_board_certification_dropdown(output_file: str):
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Provider']
    header_row = [cell.value for cell in ws[1]]
    max_row = ws.max_row
    # Build the list of ranges to clear
    ranges_to_clear = []
    for cert_num in range(1, 6):
        for col_name in [f'Board Certification {cert_num}', f'Sub Board Certification {cert_num}']:
            try:
                col_idx = header_row.index(col_name) + 1  # 1-based index
            except ValueError:
                continue
            col_letter = get_column_letter(col_idx)
            dv_range = f"{col_letter}2:{col_letter}{max_row}"
            ranges_to_clear.append(dv_range)
    # Remove data validations that overlap with these ranges
    dvs = list(ws.data_validations.dataValidation)
    for dv in dvs:
        for rng in list(dv.ranges):
            for clear_rng in ranges_to_clear:
                if str(rng) == clear_rng:
                    dv.ranges.remove(rng)
        if len(dv.ranges.ranges) == 0:
            ws.data_validations.dataValidation.remove(dv)
    # Apply new Board Certification dropdowns
    for cert_num in range(1, 6):
        col_name = f'Board Certification {cert_num}'
        try:
            col_idx = header_row.index(col_name) + 1  # 1-based index
        except ValueError:
            continue  # Skip if the column is not found
        dv = DataValidation(type="list", formula1='=ValidationAndReference!$N$2:$N$299', allow_blank=True)
        col_letter = get_column_letter(col_idx)
        dv_range = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(dv_range)
        ws.add_data_validation(dv)
    # Apply new Sub Board Certification dropdowns
    for cert_num in range(1, 6):
        col_name = f'Sub Board Certification {cert_num}'
        try:
            col_idx = header_row.index(col_name) + 1  # 1-based index
        except ValueError:
            continue  # Skip if the column is not found
        dv = DataValidation(type="list", formula1='=ValidationAndReference!$AB$2:$AB$156', allow_blank=True)
        col_letter = get_column_letter(col_idx)
        dv_range = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(dv_range)
        ws.add_data_validation(dv)
    wb.save(output_file) 