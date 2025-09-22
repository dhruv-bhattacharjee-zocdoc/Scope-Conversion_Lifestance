import openpyxl
import os
from openpyxl import Workbook

def create_npi_specialty_excel(input_excel_path, output_excel_path):
    # Load the workbook and select the first sheet
    wb = openpyxl.load_workbook(input_excel_path, data_only=True)
    sheet = wb.active

    # Find the column index for 'NPI'
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    try:
        npi_col_idx = header.index('NPI')
    except ValueError:
        raise Exception("'NPI' column not found in the Excel file.")

    # Extract all values from the 'NPI' column (skip header)
    npi_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        npi_value = row[npi_col_idx]
        if npi_value is not None:
            npi_list.append(npi_value)

    # Write the NPI list to a new Excel file
    wb_npi = Workbook()
    sheet_npi = wb_npi.active
    sheet_npi.title = 'NPI'
    # Write header
    sheet_npi['A1'] = 'NPI'
    # Write NPI values
    for idx, npi in enumerate(npi_list, start=2):
        sheet_npi[f'A{idx}'] = npi
    try:
        wb_npi.save(output_excel_path)
        print(f"Extracted {len(npi_list)} NPI values and saved to {output_excel_path}")
    except Exception as e:
        print(f"Failed to save file at {output_excel_path}. Error: {e}")
        print(f"Current working directory: {os.getcwd()}")
