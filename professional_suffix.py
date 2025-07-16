import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os
import re
from difflib import get_close_matches

def normalize_suffix(s):
    # Remove punctuation, spaces, lowercase everything
    return re.sub(r'[^a-zA-Z0-9]', '', s or '').lower()

def get_dropdown_suffixes(template_file):
    wb = openpyxl.load_workbook(template_file, data_only=True)
    ws = wb['ValidationAndReference']
    # Column G = 7, rows 2-511
    suffixes = [ws.cell(row=i, column=7).value for i in range(2, 512)]
    suffixes = [s for s in suffixes if s and str(s).strip()]
    return suffixes

def extract_professional_suffix(input_file, template_file='Excel Files/New Business Scope Sheet - Practice Locations and Providers.xlsx'):
    dropdown_suffixes = get_dropdown_suffixes(template_file)
    norm_dropdown = {normalize_suffix(s): s for s in dropdown_suffixes}
    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in.active
    if ws_in is None:
        raise ValueError("Input worksheet could not be loaded.")
    input_header_row = [cell.value for cell in ws_in[1]]
    try:
        license_idx = input_header_row.index('License Type')
    except ValueError:
        raise ValueError("'License Type' column not found in input file.")
    suffix_lists = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        cell_value = row[license_idx]
        if cell_value is not None:
            split_values = [v.strip() for v in re.split(r'[\s,]+', str(cell_value)) if v.strip()]
        else:
            split_values = [""]
        mapped = []
        for val in split_values:
            norm_val = normalize_suffix(val)
            # Try exact normalization match
            if norm_val in norm_dropdown:
                mapped.append(norm_dropdown[norm_val])
            else:
                # Fuzzy match: get closest match from dropdowns
                close = get_close_matches(norm_val, norm_dropdown.keys(), n=1, cutoff=0.8)
                if close:
                    mapped.append(norm_dropdown[close[0]])
                else:
                    mapped.append(val)
        suffix_lists.append(mapped)
    return suffix_lists


def add_professional_suffix_dropdowns(output_file):
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Provider']
    header_row = [cell.value for cell in ws[1]]
    for col_name in [f'Professional Suffix {i}' for i in range(1, 4)]:
        try:
            col_idx = header_row.index(col_name) + 1  # 1-based index for openpyxl
        except ValueError:
            continue  # Skip if the column is not found
        dv = DataValidation(type="list", formula1='=ValidationAndReference!$G$2:$G$511', allow_blank=True)
        max_row = ws.max_row
        col_letter = get_column_letter(col_idx)
        dv_range = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(dv_range)
        ws.add_data_validation(dv)
    wb.save(output_file)

if __name__ == "__main__":
    output_file = os.path.join("Excel Files", "Output.xlsx")
    add_professional_suffix_dropdowns(output_file)
