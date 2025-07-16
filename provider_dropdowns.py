import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from typing import List, Tuple

def apply_provider_dropdowns(output_file: str, dropdown_specs: List[Tuple[str, str]]):
    """
    Applies data validation dropdowns to specified columns in the Provider sheet of the output file.
    :param output_file: Path to the Excel file
    :param dropdown_specs: List of (column_name, validation_formula) tuples
    """
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Provider']
    header_row = [cell.value for cell in ws[1]]
    max_row = ws.max_row
    for col_name, formula in dropdown_specs:
        try:
            col_idx = header_row.index(col_name) + 1  # 1-based index
        except ValueError:
            continue  # Skip if the column is not found
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv_range = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(dv_range)
        ws.add_data_validation(dv)
    wb.save(output_file)

def apply_provider_formulas(output_file: str, formula_specs: List[Tuple[str, str]]):
    """
    Applies formulas to specified columns in the Provider sheet of the output file.
    :param output_file: Path to the Excel file
    :param formula_specs: List of (column_name, formula_template) tuples. Use {row} as a placeholder for the row number.
    """
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Provider']
    header_row = [cell.value for cell in ws[1]]
    max_row = ws.max_row
    for col_name, formula_template in formula_specs:
        try:
            col_idx = header_row.index(col_name) + 1  # 1-based index
        except ValueError:
            continue  # Skip if the column is not found
        for row in range(2, max_row + 1):
            formula = formula_template.replace('{row}', str(row))
            ws.cell(row=row, column=col_idx, value=formula)
    wb.save(output_file)

if __name__ == "__main__":
    # Example usage with all specified columns
    output_file = "Excel Files/Output.xlsx"
    dropdown_specs = []
    # Specialty 1-5
    for i in range(1, 6):
        dropdown_specs.append((f"Specialty {i}", "=ValidationAndReference!$K$2:$K$311"))
    # Sub Board Certification 1-5
    for i in range(1, 6):
        dropdown_specs.append((f"Sub Board Certification {i}", "=ValidationAndReference!$N$2:$N$294"))
    # Additional Languages Spoken 1-5
    for i in range(1, 6):
        dropdown_specs.append((f"Additional Languages Spoken {i}", "=ValidationAndReference!$W$2:$W$144"))
    # Enterprise Scheduling Flag (only 'Yes')
    dropdown_specs.append(("Enterprise Scheduling Flag", '"Yes"'))
    # Practice Name validation
    dropdown_specs.append(("Practice Name", "=Location!$A$8:$A$66"))
    apply_provider_dropdowns(output_file, dropdown_specs)

    # Example formula usage
    formula_specs = [
        ("Opt Out of Ratings", '=IFERROR(INDEX(ValidationAndReference!P:P,MATCH(BD{row},ValidationAndReference!Q:Q,0)),"")'),
        # Add more (column_name, formula_template) pairs as needed
    ]
    apply_provider_formulas(output_file, formula_specs) 