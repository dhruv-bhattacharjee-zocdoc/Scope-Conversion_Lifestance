import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import re

# Define the source file and the new destination filename
source_file = os.path.join('Excel Files', 'Mergedoutput.xlsx')
input_file = os.path.join('Excel Files', 'Input.xlsx')
practice_file = os.path.join('Excel Files', 'Practice-Location.xlsx')
report_file = os.path.join('Excel Files', 'Report.xlsx')
check2_file = os.path.join('Excel Files', 'check_2 sheet.xlsx')

def update_formula_row(formula, row_num):
    # This function increases all variable row refs for formulas intended to be dragged down. E.g., A2 -> A{row_num}
    if not formula or not isinstance(formula, str):
        return formula
    def repl(match):
        col = match.group(1)
        return f"{col}{row_num}"
    # Replace standalone cell refs with updated row number (skip ranges like A:A)
    return re.sub(r'([A-Z]{1,3})(2)(?!:)', repl, formula)

def copy_sheet_with_formatting(src_ws, dest_wb, dest_title, add_concat_col=False):
    if dest_title in dest_wb.sheetnames:
        std = dest_wb[dest_title]
        dest_wb.remove(std)
    dest_ws = dest_wb.create_sheet(dest_title)
    row_count = src_ws.max_row
    col_count = src_ws.max_column
    blue_fill = PatternFill(start_color='000AD6', end_color='000AD6', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)
    for i, row in enumerate(src_ws.iter_rows(max_row=row_count, max_col=col_count)):
        for j, cell in enumerate(row):
            new_cell = dest_ws.cell(row=i+1, column=j+1, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()
        # Copy row height
        if src_ws.row_dimensions[i+1].height:
            dest_ws.row_dimensions[i+1].height = src_ws.row_dimensions[i+1].height
        # For Practice ID: Add the concatenated column
        if add_concat_col:
            concat_col_letter = get_column_letter(col_count + 1)
            if i == 0:
                header_cell = dest_ws[f"{concat_col_letter}1"]
                header_cell.value = 'Add 1 and 2'
                header_cell.fill = blue_fill
                header_cell.font = white_font
            else:
                e_val = row[4].value if len(row) > 4 and row[4].value is not None else ''
                f_val = row[5].value if len(row) > 5 and row[5].value is not None else ''
                dest_ws[f"{concat_col_letter}{i+1}"] = f"{e_val} {f_val}".strip()
    # Copy column widths
    for col in src_ws.column_dimensions:
        if src_ws.column_dimensions[col].width:
            dest_ws.column_dimensions[col].width = src_ws.column_dimensions[col].width
    # Set new col width reasonable default
    if add_concat_col:
        concat_col_letter = get_column_letter(col_count + 1)
        dest_ws.column_dimensions[concat_col_letter].width = 18

def copy_header_only(src_ws, dest_wb, dest_title, formulas=None):
    if dest_title in dest_wb.sheetnames:
        std = dest_wb[dest_title]
        dest_wb.remove(std)
    dest_ws = dest_wb.create_sheet(dest_title)
    col_count = src_ws.max_column
    # Find last row with data in column A (starting from row 2)
    row_count = 1
    for i in range(2, src_ws.max_row+1):
        val = src_ws.cell(row=i, column=1).value
        if val is not None and str(val).strip() != '':
            row_count = i
    # Copy headers (row 1) value, fill, font, border, and width
    for j in range(1, col_count+1):
        src_cell = src_ws.cell(row=1, column=j)
        dest_cell = dest_ws.cell(row=1, column=j, value=src_cell.value)
        if src_cell.has_style:
            dest_cell.font = src_cell.font.copy()
            dest_cell.border = src_cell.border.copy()
            dest_cell.fill = src_cell.fill.copy()
            dest_cell.number_format = src_cell.number_format
            dest_cell.protection = src_cell.protection.copy()
            dest_cell.alignment = src_cell.alignment.copy()
        col_letter = get_column_letter(j)
        if src_ws.column_dimensions[col_letter].width:
            dest_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width
    # Row height for header if set
    if src_ws.row_dimensions[1].height:
        dest_ws.row_dimensions[1].height = src_ws.row_dimensions[1].height
    # Write formulas to all rows (starting from 2 up to last data row only)
    if formulas is not None and row_count > 1:
        for idx, formula in enumerate(formulas, start=1):
            if formula != '' and formula is not None:
                for row in range(2, row_count + 1):
                    row_formula = update_formula_row(formula, row)
                    dest_ws.cell(row=row, column=idx, value=row_formula)

if os.path.exists(source_file):
    base, ext = os.path.splitext(source_file)
    dest_file = base + '_transposed' + ext
    # Copy base file to destination
    shutil.copyfile(source_file, dest_file)
    print(f"Created copy: {dest_file}")

    dest_wb = load_workbook(dest_file)

    # Copy 'Raw' from Input.xlsx
    input_wb = load_workbook(input_file)
    input_first_sheet = input_wb.worksheets[0]
    copy_sheet_with_formatting(input_first_sheet, dest_wb, 'Raw')
    print("Added 'Raw' sheet from Input.xlsx with formatting.")

    # Copy 'Practice ID' from Practice-Location.xlsx, add 'Add 1 and 2' col
    practice_wb = load_workbook(practice_file)
    practice_first_sheet = practice_wb.worksheets[0]
    copy_sheet_with_formatting(practice_first_sheet, dest_wb, 'Practice ID', add_concat_col=True)
    print("Added 'Practice ID' sheet from Practice-Location.xlsx with formatting and 'Add 1 and 2' column.")

    # Copy 'Report' from Report.xlsx (full formatting)
    report_wb = load_workbook(report_file)
    report_first_sheet = report_wb.worksheets[0]
    copy_sheet_with_formatting(report_first_sheet, dest_wb, 'Report')
    print("Added 'Report' sheet from Report.xlsx with formatting.")

    # Copy header row from check_2 sheet.xlsx (header and col widths/colors) & insert formulas to data rows only
    check2_wb = load_workbook(check2_file)
    check2_first_sheet = check2_wb.worksheets[0]
    check2_formulas = [
        '=Provider!S2',
        '=IF(COUNTIF(Provider!S:S,A2)>1,"Duplicates found","")',
        '',  # <Empty column>
        "=XLOOKUP(Provider!BC2,'Practice ID'!B:B,'Practice ID'!A:A)",
        '=XLOOKUP(Provider!S2,Raw!I:I,Raw!E:E)',
        '=IF(TRIM(TEXT(D2,"@"))=TRIM(TEXT(E2,"@"))," ","No")',
        '',  # <Empty column>
        '=XLOOKUP(Provider!S2,Raw!I:I,Raw!Q:Q)',
        '=XLOOKUP(Provider!S2,Raw!I:I,Raw!Z:Z)',
        '',  # <Empty column>
        "=XLOOKUP(Provider!BR2,'Practice ID'!K:K,'Practice ID'!R:R,\" \")",
        "=XLOOKUP(Provider!BR2,'Practice ID'!K:K,'Practice ID'!A:A,\" \")",
        '=IF(TRIM(TEXT(L2,"@"))=TRIM(TEXT(E2,"@"))," ","No")',
        "=XLOOKUP(Provider!BR2,'Practice ID'!K:K,'Practice ID'!D:D,\" \")",
        '',  # <Empty column>
        "=XLOOKUP(Provider!BS2,'Practice ID'!K:K,'Practice ID'!R:R,\" \")",
        "=XLOOKUP(Provider!BS2,'Practice ID'!K:K,'Practice ID'!A:A,\" \")",
        '=IF(TRIM(TEXT(Q2,"@"))=TRIM(TEXT(E2,"@"))," ","No")',
        "=XLOOKUP(Provider!BS2,'Practice ID'!K:K,'Practice ID'!D:D,\" \")"
    ]
    copy_header_only(check2_first_sheet, dest_wb, 'Check 2', formulas=check2_formulas)
    print("Added 'Check 2' sheet from check_2 sheet.xlsx with header, column widths/colors, and formulas only in data rows.")

    dest_wb.save(dest_file)
else:
    print(f"Source file not found: {source_file}")
